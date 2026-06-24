#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Split integration tests chunk 36."""

from tests.integration_base import *


class TestIntegration36(IntegrationTestBase):
    def test_invoice_request_api_create_list_and_detail(self):
        import server.db as db_module
        db = db_module.get_db()
        owner_id = create_owner(db, '电子票据业主', '13800139001')
        room_id = create_room(db, building='INVAPI', room_number='3901', owner_id=owner_id)
        bill_id = create_bill(db, room_id=room_id, owner_id=owner_id, amount=88.0, status='paid', period='2029-12')
        payment_id = create_payment(db, bill_id=bill_id, amount=88.0, method='mock', operator='api-test')
        db.close()
        try:
            status, body, _ = http_post('/api/v1/invoice-requests', {
                'bill_id': str(bill_id),
                'buyer_name': '接口测试公司',
                'buyer_tax_id': 'TAX-API-001',
                'idempotency_key': 'api-invoice-request-key',
            }, self.cookie, TEST_PORT)
            self.assertEqual(status, 200)
            data = json.loads(body)['data']
            self.assertEqual(data['status'], 'pending')
            self.assertEqual(data['amount'], '88.0')
            self.assertTrue(data['request_no'].startswith('IR'))

            status, body = http_get('/api/v1/invoice-requests?status=pending', self.cookie, TEST_PORT)
            self.assertEqual(status, 200)
            items = json.loads(body)['data']['items']
            self.assertTrue(any(item['request_no'] == data['request_no'] for item in items))

            status, body = http_get('/api/v1/invoice-requests/' + data['request_no'], self.cookie, TEST_PORT)
            self.assertEqual(status, 200)
            self.assertEqual(json.loads(body)['data']['buyer_name'], '接口测试公司')
        finally:
            db = db_module.get_db()
            db.execute('DELETE FROM invoice_requests WHERE bill_id=?', (bill_id,))
            db.execute('DELETE FROM payments WHERE bill_id=?', (bill_id,))
            db.execute('DELETE FROM bills WHERE id=?', (bill_id,))
            db.execute('DELETE FROM rooms WHERE id=?', (room_id,))
            db.execute('DELETE FROM owners WHERE id=?', (owner_id,))
            db.commit(); db.close()

    def test_invoice_request_api_create_rejects_frontdesk_user(self):
        import server.db as db_module
        db = db_module.get_db()
        owner_id = create_owner(db, '客服票据业主', '13800139002')
        room_id = create_room(db, building='INVFD', room_number='3902', owner_id=owner_id)
        bill_id = create_bill(db, room_id=room_id, owner_id=owner_id, amount=66.0, status='paid', period='2029-11')
        create_payment(db, bill_id=bill_id, amount=66.0, method='mock', operator='api-test')
        db.close()
        frontdesk_cookie = self._create_user_and_login(
            'frontdesk_api_invoice', 'frontdesk123', 'frontdesk', '客服票据API'
        )
        try:
            status, body, _ = http_post('/api/v1/invoice-requests', {
                'bill_id': str(bill_id),
                'buyer_name': '客服绕过测试公司',
                'buyer_tax_id': 'TAX-FD-001',
            }, frontdesk_cookie, TEST_PORT)
            self.assertEqual(status, 403)
            payload = json.loads(body)
            self.assertFalse(payload['ok'])
            self.assertEqual(payload['error']['code'], 'forbidden')

            db = db_module.get_db()
            count = db.execute('SELECT COUNT(*) FROM invoice_requests WHERE bill_id=?', (bill_id,)).fetchone()[0]
            db.close()
            self.assertEqual(count, 0)
        finally:
            db = db_module.get_db()
            db.execute('DELETE FROM invoice_requests WHERE bill_id=?', (bill_id,))
            db.execute('DELETE FROM payments WHERE bill_id=?', (bill_id,))
            db.execute('DELETE FROM bills WHERE id=?', (bill_id,))
            db.execute('DELETE FROM rooms WHERE id=?', (room_id,))
            db.execute('DELETE FROM owners WHERE id=?', (owner_id,))
            db.commit(); db.close()



    def _seed_reset_contract_fixture(self, suffix):
        import server.db as db_module
        db = db_module.get_db()
        fee_id = db.execute(
            "INSERT INTO fee_types(name,calc_method,unit_price) VALUES(?,?,?)",
            (f'清空范围收费项{suffix}', 'fixed', 12),
        ).lastrowid
        owner_id = db.execute(
            "INSERT INTO owners(name,phone) VALUES(?,?)",
            (f'清空范围业主{suffix}', '13800139999'),
        ).lastrowid
        room_id = db.execute(
            "INSERT INTO rooms(building,unit,room_number,floor,category,area,owner_id) VALUES(?,?,?,?,?,?,?)",
            ('RESET-SCOPE', 'A座', suffix, 1, '商户', 10, owner_id),
        ).lastrowid
        space_id = db.execute(
            "INSERT INTO commercial_spaces(space_no,floor,area,shop_name,merchant_name) VALUES(?,?,?,?,?)",
            (f'RESET-SP-{suffix}', 1, 10, f'范围店铺{suffix}', f'范围商户{suffix}'),
        ).lastrowid
        contract_id = db.execute(
            """INSERT INTO merchant_contracts(room_id,commercial_space_id,owner_id,contract_no,merchant_name,shop_name,
               rent_amount,rent_cycle,property_rate,property_cycle,deposit_amount,contract_area,building_area,start_date,end_date,status)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (room_id, space_id, owner_id, f'RESET-HT-{suffix}', f'范围商户{suffix}', f'范围店铺{suffix}', 1000,
             'monthly', 5, 'monthly', 1000, 10, 12, '2026-01-01', '2026-12-31', 'active'),
        ).lastrowid
        bill_id = db.execute(
            "INSERT INTO bills(room_id,owner_id,fee_type_id,billing_period,amount,status,bill_number) VALUES(?,?,?,?,?,?,?)",
            (room_id, owner_id, fee_id, '2031-01', 12, 'unpaid', f'RESET-SCOPE-BILL-{suffix}'),
        ).lastrowid
        db.execute("INSERT INTO payments(bill_id,amount_paid,payment_method,operator) VALUES(?,?,?,?)", (bill_id, 5, 'cash', 'tester'))
        db.execute(
            "INSERT INTO contract_bill_runs(contract_id,billing_period,generated_count,total_amount,operator) VALUES(?,?,?,?,?)",
            (contract_id, '2031-01', 1, 100, 'tester'),
        )
        db.execute(
            """INSERT INTO contract_attachments(contract_id,attachment_type,original_name,stored_name,file_ext,mime_type,file_size,uploaded_by)
               VALUES(?,?,?,?,?,?,?,?)""",
            (contract_id, '合同扫描件', 'reset.pdf', 'reset-stored.pdf', '.pdf', 'application/pdf', 8, 'admin'),
        )
        db.commit(); db.close()
        return fee_id, owner_id, room_id, space_id, contract_id, bill_id

    def test_trial_data_reset_default_keeps_contract_scope(self):
        import server.db as db_module
        self._seed_reset_contract_fixture('KEEP')
        status, body = http_get('/trial_data_reset', self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        self.assertIn('清空范围', body)
        self.assertIn('基础业务资料和收费记录', body)
        self.assertIn('合同档案和商业空间', body)
        self.assertIn('合同附件', body)

        status, _, loc = http_post('/trial_data_reset', {'confirm_text': 'RESET'}, self.cookie, TEST_PORT)
        self.assertEqual(status, 302)
        self.assertIn('/trial_data_reset', loc)
        db = db_module.get_db()
        self.assertEqual(db.execute('SELECT COUNT(*) FROM owners').fetchone()[0], 0)
        self.assertEqual(db.execute('SELECT COUNT(*) FROM rooms').fetchone()[0], 0)
        self.assertGreaterEqual(db.execute("SELECT COUNT(*) FROM merchant_contracts WHERE contract_no='RESET-HT-KEEP'").fetchone()[0], 1)
        self.assertGreaterEqual(db.execute("SELECT COUNT(*) FROM commercial_spaces WHERE space_no='RESET-SP-KEEP'").fetchone()[0], 1)
        self.assertGreaterEqual(db.execute('SELECT COUNT(*) FROM contract_bill_runs').fetchone()[0], 1)
        self.assertGreaterEqual(db.execute('SELECT COUNT(*) FROM contract_attachments').fetchone()[0], 1)
        db.close()

    def test_trial_data_reset_can_include_contract_scope(self):
        import server.db as db_module
        self._seed_reset_contract_fixture('DELETE')
        status, _, loc = http_post('/trial_data_reset', {
            'scope_form': '1',
            'confirm_text': 'RESET',
            'include_contracts': 'on',
        }, self.cookie, TEST_PORT)
        self.assertEqual(status, 302)
        self.assertIn('合同档案', urllib.parse.unquote(loc))
        db = db_module.get_db()
        self.assertGreaterEqual(db.execute("SELECT COUNT(*) FROM owners WHERE name='清空范围业主DELETE'").fetchone()[0], 1)
        self.assertGreaterEqual(db.execute("SELECT COUNT(*) FROM rooms WHERE room_number='DELETE'").fetchone()[0], 1)
        self.assertEqual(db.execute("SELECT COUNT(*) FROM merchant_contracts WHERE contract_no='RESET-HT-DELETE'").fetchone()[0], 0)
        self.assertEqual(db.execute("SELECT COUNT(*) FROM commercial_spaces WHERE space_no='RESET-SP-DELETE'").fetchone()[0], 0)
        self.assertEqual(db.execute('SELECT COUNT(*) FROM contract_bill_runs').fetchone()[0], 0)
        self.assertEqual(db.execute('SELECT COUNT(*) FROM contract_attachments').fetchone()[0], 0)
        db.close()


    def test_trial_data_reset_scope_form_requires_one_scope(self):
        status, _, loc = http_post('/trial_data_reset', {
            'scope_form': '1',
            'confirm_text': 'RESET',
        }, self.cookie, TEST_PORT)
        self.assertEqual(status, 302)
        self.assertIn('请至少选择一个清空范围', urllib.parse.unquote(loc))

    def test_auto_detect_basic_room_file_with_b_tower_room_number_uses_room_preview(self):
        boundary = '----PropertyAutoDetectRoomBoundary'
        csv_data = (
            '楼栋,单元/座,铺位号/房号,房间类型,面积㎡,业主姓名,租户姓名,合同开始日期,合同结束日期,缴费周期\n'
            'B座,B座,1601-16,住宅,66.5,自动检测业主,自动检测租户,2026-01-01,2026-12-31,月付\n'
        )
        body = (
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="data_type"\r\n\r\nauto\r\n'
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="mode"\r\n\r\npreview\r\n'
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="file"; filename="auto-basic-rooms.csv"\r\n'
            'Content-Type: text/csv\r\n\r\n'
            f'{csv_data}\r\n'
            f'--{boundary}--\r\n'
        ).encode('utf-8')
        conn = http.client.HTTPConnection(BASE_URL, TEST_PORT)
        conn.request('POST', '/import/upload', body, {
            'Content-Type': f'multipart/form-data; boundary={boundary}',
            'Content-Length': str(len(body)),
            'Cookie': self.cookie,
        })
        resp = conn.getresponse()
        html = resp.read().decode('utf-8', errors='ignore')
        conn.close()
        self.assertEqual(resp.status, 200)
        self.assertIn('数据核对与修正', html)
        self.assertIn('房间信息', html)
        self.assertNotIn('B座出租合同导入预览', html)
        self.assertIn('1601-16', html)

    def test_import_page_offers_templates_for_each_import_type_and_auto_warning(self):
        status, body = http_get('/import', self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        for href in [
            '/import/template/basic.xlsx',
            '/import/template/owners.xlsx',
            '/import/template/payment_ledger.xlsx',
            '/import/template/bills.xlsx',
            '/import/template/commercial_contracts.xlsx',
            '/import/template/b_tower_contracts.xlsx',
        ]:
            self.assertIn(href, body)
        self.assertIn('自动检测适合模板文件或字段清晰的 Excel', body)
        self.assertIn('识别不准确', body)

    def test_import_type_templates_download_with_expected_headers(self):
        import io
        import openpyxl
        expected = {
            '/import/template/owners.xlsx': ('业主信息模板', ['业主姓名', '业主电话', '身份证号', '备注']),
            '/import/template/payment_ledger.xlsx': ('收款明细识别模板', ['缴费日期', '房间号', '用户名称', '本期金额', '本期收款(合计)', '物业费']),
            '/import/template/bills.xlsx': ('账单记录模板', ['楼栋', '房号', '费用项目', '账期', '金额', '状态']),
            '/import/template/b_tower_contracts.xlsx': ('B座出租合同模板', ['楼栋', '单元/座', '房号', '租户', '租户电话', '面积', '物业费单价', '缴费周期', '合同开始日期', '合同结束日期']),
        }
        for path, (sheet_name, headers_prefix) in expected.items():
            conn = http.client.HTTPConnection(BASE_URL, TEST_PORT)
            conn.request('GET', path, headers={'Cookie': self.cookie})
            resp = conn.getresponse()
            data = resp.read()
            content_type = resp.getheader('Content-Type', '')
            conn.close()
            self.assertEqual(resp.status, 200, path)
            self.assertIn('spreadsheetml.sheet', content_type)
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            self.assertIn(sheet_name, wb.sheetnames)
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]][:len(headers_prefix)]
            self.assertEqual(headers, headers_prefix)
            wb.close()

    def test_commercial_contract_template_uses_required_rented_contract_sheet(self):
        import io
        import openpyxl
        conn = http.client.HTTPConnection(BASE_URL, TEST_PORT)
        conn.request('GET', '/import/template/commercial_contracts.xlsx', headers={'Cookie': self.cookie})
        resp = conn.getresponse()
        data = resp.read()
        conn.close()
        self.assertEqual(resp.status, 200)
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        self.assertIn('在租合同', wb.sheetnames)
        headers = [cell.value for cell in wb['在租合同'][1]]
        for name in ['楼层', '合同编号', '商铺号', '起租期（日）', '承租人', '合同面积（㎡）', '租赁期限（年）']:
            self.assertIn(name, headers)
        wb.close()
