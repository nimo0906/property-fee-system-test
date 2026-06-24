#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Split integration tests chunk 28."""

from tests.integration_base import *


class TestIntegration28(IntegrationTestBase):
    def test_reports_reconciliation_print_page_uses_current_filters(self):
        import server.db as db_module
        db = db_module.get_db()
        owner_id = db.execute("INSERT INTO owners(name, phone) VALUES('打印对账业主', '12900000000')").lastrowid
        room_a = db.execute("""
            INSERT INTO rooms(building, unit, room_number, floor, category, area, owner_id)
            VALUES('PRINTA', 'A座', '1801', 18, '居民', 70, ?)
        """, (owner_id,)).lastrowid
        room_b = db.execute("""
            INSERT INTO rooms(building, unit, room_number, floor, category, area, owner_id)
            VALUES('PRINTB', 'A座', '1802', 18, '居民', 70, ?)
        """, (owner_id,)).lastrowid
        db.execute("""
            INSERT INTO bills(room_id, owner_id, fee_type_id, billing_period, amount, due_date, status, bill_number)
            VALUES(?, ?, 1, '2030-10', 120, '2030-10-28', 'unpaid', 'PRINT-A-UNPAID')
        """, (room_a, owner_id))
        db.execute("""
            INSERT INTO bills(room_id, owner_id, fee_type_id, billing_period, amount, due_date, status, bill_number)
            VALUES(?, ?, 1, '2030-10', 90, '2030-10-28', 'paid', 'PRINT-B-PAID')
        """, (room_b, owner_id))
        db.commit()
        db.close()

        status, html = http_get('/reports?period=2030-10&building=PRINTA&status=unpaid', self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        self.assertIn('/reports/reconciliation/print?period_start=2030-10-01&amp;period_end=2030-10-31&amp;building=PRINTA&amp;status=unpaid', html)

        status, print_html = http_get('/reports/reconciliation/print?period=2030-10&building=PRINTA&status=unpaid', self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        self.assertIn('对账单打印', print_html)
        self.assertIn('2030-10', print_html)
        self.assertIn('PRINTA', print_html)
        self.assertIn('未缴', print_html)
        self.assertIn('PRINT-A-UNPAID', print_html)
        self.assertNotIn('PRINT-B-PAID', print_html)
        self.assertIn('window.print()', print_html)
        self.assertIn('@media print', print_html)
        self.assertIn('class="print-toolbar"', print_html)
        self.assertIn('保存为PDF', print_html)
        self.assertIn('账单状态分布', print_html)
        self.assertIn('未缴账单', print_html)
        self.assertIn('本期回款缺口', print_html)


    def test_reports_reconciliation_filters_by_building_and_status(self):
        import server.db as db_module
        db = db_module.get_db()
        owner_id = db.execute("INSERT INTO owners(name, phone) VALUES('筛选对账业主', '13000000000')").lastrowid
        room_a = db.execute("""
            INSERT INTO rooms(building, unit, room_number, floor, category, area, owner_id)
            VALUES('FILTERA', 'A座', '1701', 17, '居民', 70, ?)
        """, (owner_id,)).lastrowid
        room_b = db.execute("""
            INSERT INTO rooms(building, unit, room_number, floor, category, area, owner_id)
            VALUES('FILTERB', 'A座', '1702', 17, '居民', 70, ?)
        """, (owner_id,)).lastrowid
        db.execute("""
            INSERT INTO bills(room_id, owner_id, fee_type_id, billing_period, amount, due_date, status, bill_number)
            VALUES(?, ?, 1, '2030-09', 70, '2030-09-28', 'unpaid', 'FILTER-A-UNPAID')
        """, (room_a, owner_id))
        paid_id = db.execute("""
            INSERT INTO bills(room_id, owner_id, fee_type_id, billing_period, amount, due_date, status, bill_number)
            VALUES(?, ?, 1, '2030-09', 90, '2030-09-28', 'paid', 'FILTER-B-PAID')
        """, (room_b, owner_id)).lastrowid
        db.execute("INSERT INTO payments(bill_id, amount_paid, payment_method, operator) VALUES(?, 90, 'cash', '筛选员')", (paid_id,))
        db.commit()
        db.close()

        status, html = http_get('/reports?period=2030-09&building=FILTERA&status=unpaid', self.cookie, TEST_PORT)

        self.assertEqual(status, 200)
        self.assertIn('FILTERA', html)
        self.assertIn('FILTER-A-UNPAID', html)
        self.assertNotIn('FILTER-B-PAID', html)
        self.assertIn('name="status"', html)
        self.assertIn('/reports/reconciliation.csv?period_start=2030-09-01&amp;period_end=2030-09-30&amp;building=FILTERA&amp;status=unpaid', html)

        status, csv = http_get('/reports/reconciliation.csv?period=2030-09&building=FILTERA&status=unpaid', self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        self.assertIn('FILTER-A-UNPAID', csv)
        self.assertNotIn('FILTER-B-PAID', csv)


    def test_reports_reconciliation_csv_export(self):
        import server.db as db_module
        db = db_module.get_db()
        owner_id = db.execute("INSERT INTO owners(name, phone) VALUES('CSV对账业主', '13100000000')").lastrowid
        room_id = db.execute("""
            INSERT INTO rooms(building, unit, room_number, floor, category, area, owner_id)
            VALUES('REPORTCSV', 'A座', '1601', 16, '居民', 75, ?)
        """, (owner_id,)).lastrowid
        db.execute("""
            INSERT INTO bills(room_id, owner_id, fee_type_id, billing_period, amount, due_date, status, bill_number)
            VALUES(?, ?, 1, '2030-08', 88, '2030-08-28', 'unpaid', 'RPT-CSV')
        """, (room_id, owner_id))
        db.commit()
        db.close()

        status, csv = http_get('/reports/reconciliation.csv?period=2030-08', self.cookie, TEST_PORT)

        self.assertEqual(status, 200)
        self.assertIn('building,room,owner,phone,fee_type,period,amount,paid,due,status,due_date,bill_number', csv)
        self.assertIn('REPORTCSV', csv)
        self.assertIn('RPT-CSV', csv)
        self.assertIn('88.0', csv)


    def test_reports_page(self):
        _, body = http_get('/reports?period=2026-06', self.cookie, TEST_PORT)
        self.assertIn('对账报表', body)
        self.assertIn('应收总额', body)


    def test_reports_show_waiver_detail_for_adjusted_bills(self):
        import server.db as db_module
        db = db_module.get_db()
        owner_id = db.execute("INSERT INTO owners(name, phone) VALUES('减免明细业主', '13200000000')").lastrowid
        room_id = db.execute("""
            INSERT INTO rooms(building, unit, room_number, floor, category, area, owner_id)
            VALUES('WAIVE', 'A座', '1501', 15, '居民', 80, ?)
        """, (owner_id,)).lastrowid
        bill_id = db.execute("""
            INSERT INTO bills(room_id, owner_id, fee_type_id, billing_period, amount, due_date, status, bill_number)
            VALUES(?, ?, 1, '2026-06', 100, '2026-06-28', 'partial', 'WAIVE-1501-01')
        """, (room_id, owner_id)).lastrowid
        db.execute("INSERT INTO payments(bill_id, amount_paid, payment_method, operator) VALUES(?, 27.6, 'cash', '减免测试员')", (bill_id,))
        db.execute("""
            INSERT INTO bill_adjustments(bill_id, old_amount, new_amount, reason, approved_by)
            VALUES(?, 100, 27.6, '困难减免', '财务主管')
        """, (bill_id,))
        db.commit()
        db.close()

        status, html = http_get('/reports?period=2026-06&building=WAIVE&status=partial', self.cookie, TEST_PORT)

        self.assertEqual(status, 200)
        self.assertIn('减免 ¥72.4 · 1笔', html)
        self.assertIn('href="#reportWaivers"', html)
        self.assertIn('减免明细', html)
        self.assertIn('WAIVE-1501-01', html)
        self.assertIn('减免明细业主', html)
        self.assertIn('困难减免', html)
        self.assertIn('财务主管', html)
        self.assertIn('¥100.0', html)
        self.assertIn('¥27.6', html)
        self.assertIn('¥72.4', html)


    def test_reports_kpis_link_directly_to_reconciliation_details(self):
        import server.db as db_module
        db = db_module.get_db()
        owner_id = db.execute("INSERT INTO owners(name, phone) VALUES('KPI明细业主', '13200000001')").lastrowid
        room_id = db.execute("""
            INSERT INTO rooms(building, unit, room_number, floor, category, area, owner_id)
            VALUES('KPIDETAIL', 'A座', '1601', 16, '居民', 80, ?)
        """, (owner_id,)).lastrowid
        paid_bill = db.execute("""
            INSERT INTO bills(room_id, owner_id, fee_type_id, billing_period, amount, due_date, status, bill_number)
            VALUES(?, ?, 1, '2028-12', 40, '2028-12-28', 'paid', 'KPIDETAIL-PAID')
        """, (room_id, owner_id)).lastrowid
        db.execute("INSERT INTO payments(bill_id, amount_paid, payment_method, operator) VALUES(?, 40, 'cash', 'KPI测试员')", (paid_bill,))
        db.execute("""
            INSERT INTO bills(room_id, owner_id, fee_type_id, billing_period, amount, due_date, status, bill_number)
            VALUES(?, ?, 2, '2028-12', 60, '2028-12-28', 'unpaid', 'KPIDETAIL-DUE')
        """, (room_id, owner_id))
        db.commit()
        db.close()

        status, html = http_get('/reports?period=2028-12&building=KPIDETAIL', self.cookie, TEST_PORT)

        self.assertEqual(status, 200)
        self.assertIn('href="#reportArrears"', html)
        self.assertIn('href="#reportCollection"', html)
        self.assertIn('href="#reportWaivers"', html)
        self.assertIn('KPIDETAIL-DUE', html)
        self.assertIn('KPI测试员', html)


    def test_reports_show_enterprise_analysis_sections(self):
        from server.db import get_db
        db = get_db()
        owner_id = create_owner(db, '报表经营商户', '13900007777')
        mall_room = create_room(db, building='商场', unit='商场', room_number='RPT-M1', category='商户', area=100, owner_id=owner_id)
        b_room = create_room(db, building='B座', unit='B座', room_number='RPT-B1', category='居民', area=80, owner_id=owner_id)
        db.execute("UPDATE rooms SET shop_name='报表贡献店' WHERE id=?", (mall_room,))
        mall_bill = create_bill(db, room_id=mall_room, owner_id=owner_id, fee_type_id=1, period='2032-06', amount=900, status='unpaid')
        b_bill = create_bill(db, room_id=b_room, owner_id=owner_id, fee_type_id=2, period='2032-06', amount=100, status='paid')
        create_payment(db, bill_id=b_bill, amount=100, method='cash', operator='报表经营测试')
        db.commit(); db.close()

        status, body = http_get('/reports?period=2032-06', self.cookie, TEST_PORT)

        self.assertEqual(status, 200)
        for text in ['经营分析看板', 'B座 / 商场对比', '欠费趋势', '商户贡献', '费用结构']:
            self.assertIn(text, body)
        self.assertIn('报表贡献店', body)
        self.assertIn('收缴率', body)



    def test_reports_fee_type_summary_shows_main_and_groups_low_frequency(self):
        import server.db as db_module
        db = db_module.get_db()
        owner_id = create_owner(db, '费用类型报表业主', '13900009991')
        room_id = create_room(db, building='FEEMAIN', unit='A座', room_number='901', floor=9, category='居民', area=60, owner_id=owner_id)
        main_fee = db.execute("SELECT id FROM fee_types WHERE name='物业费(居民)'").fetchone()[0]
        low_fee = db.execute("""
            INSERT INTO fee_types(name,calc_method,unit_price,unit,billing_cycle,sort_order,is_active,notes)
            VALUES('测试低频服务费','fixed',77,'元','monthly',299,1,'报表低频聚合测试')
        """).lastrowid
        db.execute("""
            INSERT INTO fee_types(name,calc_method,unit_price,unit,billing_cycle,sort_order,is_active,notes)
            VALUES('测试低频零金额项','fixed',0,'元','monthly',300,1,'报表零数据隐藏测试')
        """)
        main_bill = create_bill(db, room_id=room_id, owner_id=owner_id, fee_type_id=main_fee, period='2044-01', amount=120, status='paid')
        create_payment(db, bill_id=main_bill, amount=120, method='cash', operator='报表测试')
        create_bill(db, room_id=room_id, owner_id=owner_id, fee_type_id=low_fee, period='2044-01', amount=77, status='unpaid')
        db.commit(); db.close()

        status, html = http_get('/reports?period=2044-01&building=FEEMAIN', self.cookie, TEST_PORT)

        self.assertEqual(status, 200)
        self.assertIn('主要收费类型统计', html)
        summary_html = html.split('主要收费类型统计', 1)[1].split('收款日报', 1)[0]
        self.assertIn('物业费(居民)', summary_html)
        self.assertIn('其他收费', summary_html)
        self.assertIn('¥77.0', summary_html)
        self.assertNotIn('测试低频服务费', summary_html)
        self.assertNotIn('测试低频零金额项', html)

    def test_reports_enterprise_analysis_xlsx_export(self):
        from server.db import get_db
        db = get_db()
        owner_id = create_owner(db, '导出经营商户', '13900008888')
        mall_room = create_room(db, building='商场', unit='商场', room_number='XLS-M1', category='商户', area=120, owner_id=owner_id)
        b_room = create_room(db, building='B座', unit='B座', room_number='XLS-B1', category='居民', area=80, owner_id=owner_id)
        db.execute("UPDATE rooms SET shop_name='导出贡献店' WHERE id=?", (mall_room,))
        mall_bill = create_bill(db, room_id=mall_room, owner_id=owner_id, fee_type_id=1, period='2033-01', amount=1200, status='unpaid')
        b_bill = create_bill(db, room_id=b_room, owner_id=owner_id, fee_type_id=2, period='2033-01', amount=300, status='paid')
        create_payment(db, bill_id=b_bill, amount=300, method='cash', operator='经营导出测试')
        db.commit(); db.close()

        status, html = http_get('/reports?period=2033-01', self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        self.assertIn('/reports/enterprise_analysis.xlsx?period_start=2033-01-01&amp;period_end=2033-01-31', html)

        conn = http.client.HTTPConnection(BASE_URL, TEST_PORT)
        conn.request('GET', '/reports/enterprise_analysis.xlsx?period=2033-01', headers={'Cookie': self.cookie})
        resp = conn.getresponse()
        data = resp.read()
        content_type = resp.getheader('Content-Type', '')
        conn.close()

        self.assertEqual(resp.status, 200)
        self.assertIn('spreadsheetml.sheet', content_type)

        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        self.assertEqual(wb.sheetnames, ['经营概览', 'B座商场对比', '欠费趋势', '商户贡献', '费用结构'])
        self.assertEqual(wb['经营概览']['A1'].value, '经营分析看板')
        self.assertEqual(wb['经营概览']['B3'].value, 1500)
        self.assertEqual(wb['经营概览']['B4'].value, 300)
        self.assertEqual(wb['B座商场对比']['A2'].value, 'B座')
        self.assertIn('导出贡献店', [row[0] for row in wb['商户贡献'].iter_rows(min_row=2, values_only=True)])
        self.assertIn('物业费', ''.join(str(row[0] or '') for row in wb['费用结构'].iter_rows(min_row=2, values_only=True)))
        wb.close()


    def test_bill_export_csv(self):
        self._ensure_room_for_billing()
        http_post('/bills/generate', {
            'period': '2026-06', 'fee_type_ids': '1', 'due_day': '28',
        }, self.cookie, TEST_PORT)

        conn = http.client.HTTPConnection(BASE_URL, TEST_PORT)
        conn.request('GET', '/bills/export?period=2026-06',
                     headers={'Cookie': self.cookie})
        resp = conn.getresponse()
        data = resp.read().decode('utf-8-sig')
        ct = resp.headers.get('Content-Type', '')
        conn.close()
        self.assertEqual(resp.status, 200)
        self.assertIn('票据编号', data)
        self.assertIn('csv', ct)


    def test_commercial_fee_group_add_button_preserves_group_context(self):
        status, body = http_get('/fee_types?group=commercial', self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        self.assertIn('/fee_types/create?group=commercial', body)

        status, body = http_get('/fee_types/create?group=commercial', self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        self.assertIn('添加商业公司收费项目', body)
        self.assertIn('name="return_group" value="commercial"', body)

        status, body, loc = http_post('/fee_types/create', {
            'name': '测试商业收费项',
            'calc_method': 'fixed',
            'unit_price': '12.5',
            'unit': '元',
            'billing_cycle': 'monthly',
            'sort_order': '39',
            'is_active': 'on',
            'notes': '商业测试收费',
            'reminder_advance_days': '15',
            'return_group': 'commercial',
        }, self.cookie, TEST_PORT)
        self.assertEqual(status, 302)
        self.assertIn('/fee_types?group=commercial&flash=', loc)

        _, body = http_get('/fee_types?group=commercial', self.cookie, TEST_PORT)
        self.assertIn('测试商业收费项', body)

    def test_fee_type_edit_page_preserves_saved_select_values(self):
        import server.db as db_module
        db = db_module.get_db()
        fee_id = db.execute("""
            INSERT INTO fee_types(name,calc_method,unit_price,unit,billing_cycle,sort_order,is_active,notes,reminder_advance_days)
            VALUES('测试编辑保留选择','area',1.2,'元/m2','monthly',88,1,'编辑前',30)
        """).lastrowid
        db.commit()
        db.close()

        status, body, loc = http_post(f'/fee_types/{fee_id}/edit', {
            'name': '测试编辑保留选择',
            'calc_method': 'meter',
            'unit_price': '9.5',
            'unit': '元/吨',
            'billing_cycle': 'quarterly',
            'sort_order': '89',
            'is_active': 'on',
            'notes': '编辑后',
            'reminder_advance_days': '12',
        }, self.cookie, TEST_PORT)
        self.assertEqual(status, 302)
        self.assertIn(f'/fee_types/{fee_id}/edit', loc)

        status, body = http_get(f'/fee_types/{fee_id}/edit', self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        self.assertIn('<option value="meter" selected>按用量×单价</option>', body)
        self.assertIn('<option value="quarterly" selected>每季度</option>', body)

        db = db_module.get_db()
        row = db.execute('SELECT calc_method,billing_cycle,unit_price,unit,notes FROM fee_types WHERE id=?', (fee_id,)).fetchone()
        db.close()
        self.assertEqual(row['calc_method'], 'meter')
        self.assertEqual(row['billing_cycle'], 'quarterly')
        self.assertEqual(float(row['unit_price']), 9.5)
        self.assertEqual(row['unit'], '元/吨')
        self.assertEqual(row['notes'], '编辑后')
