#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Split integration tests chunk 33."""

from tests.integration_base import *


class TestIntegration33(IntegrationTestBase):
    def test_import_preview_shows_audit_sections_and_problem_export(self):
        boundary = '----PropertyImportPreviewAuditBoundary'
        csv_data = (
            '类别,房号,业主姓名,面积㎡,联系电话,合同缴租期\n'
            '商户,B座930,预览正常业主,88.5,13900008881,2026.1.1-2026.12.31\n'
            '商户,美容护肤,广告招商,0,,\n'
            '商户,B座931,日期异常业主,66.0,13900008882,bad-date\n'
        )
        body = (
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="data_type"\r\n\r\nrooms\r\n'
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="mode"\r\n\r\npreview\r\n'
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="file"; filename="preview-audit.csv"\r\n'
            'Content-Type: text/csv\r\n\r\n'
            f'{csv_data}\r\n'
            f'--{boundary}--\r\n'
        ).encode('utf-8')

        conn = http.client.HTTPConnection(BASE_URL, TEST_PORT)
        conn.request('POST', '/import/upload', body, {
            'Content-Type': f'multipart/form-data; boundary={boundary}',
            'Content-Length': str(len(body)),
            'Cookie': self.cookie,
        })
        resp = conn.getresponse()
        preview_html = resp.read().decode('utf-8')
        conn.close()

        self.assertEqual(resp.status, 200)
        self.assertIn('数据核对与修正', preview_html)
        self.assertIn('问题提醒', preview_html)
        self.assertIn('高级识别信息', preview_html)
        self.assertIn('确认导入并生成结果', preview_html)
        self.assertIn('预览正常业主', preview_html)
        self.assertIn('非房间资料', preview_html)
        self.assertIn('合同日期异常', preview_html)
        match = re.search(r'href="(/import/problem_rows/[^"]+\.csv)"', preview_html)
        self.assertIsNotNone(match)

        status, csv_body = http_get(match.group(1), self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        self.assertIn('美容护肤', csv_body)
        self.assertIn('bad-date', csv_body)


    def test_import_result_shows_and_exports_problem_rows(self):
        boundary = '----PropertyImportProblemRowsBoundary'
        csv_data = (
            '类别,房号,业主姓名,面积㎡,联系电话,合同缴租期\n'
            '商户,B座903,正常业主,10,13900007777,2024.1.1-2026.1.1\n'
            '商户,,无房号,10,13900007778,\n'
            '商户,业态,无效房号,10,13900007779,\n'
            '商户,B座904,日期异常,10,13900007770,bad-date\n'
        )
        body = (
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="data_type"\r\n\r\nrooms\r\n'
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="mode"\r\n\r\nimport\r\n'
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="file"; filename="problem-rows.csv"\r\n'
            'Content-Type: text/csv\r\n\r\n'
            f'{csv_data}\r\n'
            f'--{boundary}--\r\n'
        ).encode('utf-8')

        conn = http.client.HTTPConnection(BASE_URL, TEST_PORT)
        conn.request('POST', '/import/upload', body, {
            'Content-Type': f'multipart/form-data; boundary={boundary}',
            'Content-Length': str(len(body)),
            'Cookie': self.cookie,
        })
        resp = conn.getresponse()
        html = resp.read().decode('utf-8')
        conn.close()

        self.assertEqual(resp.status, 200)
        self.assertIn('问题行明细', html)
        self.assertIn('缺少房号', html)
        self.assertIn('房号无效', html)
        self.assertIn('合同日期异常', html)
        match = re.search(r'href="(/import/problem_rows/[^"]+\.csv)"', html)
        self.assertIsNotNone(match)

        status, csv_body = http_get(match.group(1), self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        self.assertIn('行号,原因,原始数据', csv_body)
        self.assertIn('缺少房号', csv_body)
        self.assertIn('房号无效', csv_body)
        self.assertIn('合同日期异常', csv_body)


    def test_import_page_downloads_basic_info_template(self):
        status, body = http_get('/import', self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        self.assertIn('下载基础资料模板', body)
        self.assertIn('/import/template/basic.xlsx', body)

        conn = http.client.HTTPConnection(BASE_URL, TEST_PORT)
        conn.request('GET', '/import/template/basic.xlsx', headers={'Cookie': self.cookie})
        resp = conn.getresponse()
        data = resp.read()
        content_type = resp.getheader('Content-Type', '')
        conn.close()
        self.assertEqual(resp.status, 200)
        self.assertIn('spreadsheetml.sheet', content_type)

        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb['基础资料导入模板']
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, [
            '楼栋', '单元/座', '铺位号/房号', '楼层',
            '房间类型', '面积㎡', '物业费单价', '水费标准',
            '业主姓名', '业主电话', '身份证号',
            '租户姓名', '租户电话', '租户身份证号',
            '合同开始日期', '合同结束日期', '缴费周期',
            '店铺名称', '业态/商户类别', '备注',
        ])
        self.assertEqual(ws['A2'].value, '示例项目')
        self.assertEqual(ws['C2'].value, '1F-101')
        self.assertIn('历史收款金额', ws['T2'].value)
        self.assertIn('填写说明', wb.sheetnames)
        wb.close()

        status, csv_body = http_get('/import/template/basic.csv', self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        self.assertIn('楼栋,单元/座,铺位号/房号,楼层,房间类型,面积㎡', csv_body)


    def test_import_page_guides_real_data_workflow(self):
        status, body = http_get('/import', self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        self.assertIn('导入向导', body)
        self.assertIn('基础资料导入', body)
        self.assertIn('收款明细识别', body)
        self.assertIn('安全机制', body)
        self.assertIn('自动备份', body)
        self.assertIn('一键撤销', body)
        self.assertIn('历史金额不会自动入账', body)


    def test_import_result_does_not_show_bill_generation_shortcut(self):
        boundary = '----PropertyImportBillingNextBoundary'
        csv_data = '楼栋,类别,房号,业主姓名,面积㎡,联系电话,合同缴租期\nNEXTBILL,商户,901,导入收费业主,88.5,13900006666,2026.1.1-2026.12.31\n'
        body = (
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="data_type"\r\n\r\nrooms\r\n'
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="mode"\r\n\r\nimport\r\n'
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="file"; filename="next-billing.csv"\r\n'
            'Content-Type: text/csv\r\n\r\n'
            f'{csv_data}\r\n'
            f'--{boundary}--\r\n'
        ).encode('utf-8')

        conn = http.client.HTTPConnection(BASE_URL, TEST_PORT)
        conn.request('POST', '/import/upload', body, {
            'Content-Type': f'multipart/form-data; boundary={boundary}',
            'Content-Length': str(len(body)),
            'Cookie': self.cookie,
        })
        resp = conn.getresponse()
        html = resp.read().decode('utf-8')
        conn.close()

        self.assertEqual(resp.status, 200)
        self.assertNotIn('下一步生成账单', html)
        self.assertNotIn('/bills/generate?building=NEXTBILL', html)
        self.assertIn('继续导入', html)
        self.assertIn('查看房间', html)


    def test_import_creates_backup_and_shows_result_details(self):
        boundary = '----PropertyImportTestBoundary'
        csv_data = '类别,房号,业主姓名,面积㎡,联系电话,合同缴租期\n商户,B座901,自动备份业主,88.5,13900009999,2024.1.1-2026.12.31\n'
        body = (
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="data_type"\r\n\r\nrooms\r\n'
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="mode"\r\n\r\nimport\r\n'
            f'--{boundary}\r\n'
            'Content-Disposition: form-data; name="file"; filename="rooms.csv"\r\n'
            'Content-Type: text/csv\r\n\r\n'
            f'{csv_data}\r\n'
            f'--{boundary}--\r\n'
        ).encode('utf-8')

        conn = http.client.HTTPConnection(BASE_URL, TEST_PORT)
        conn.request('POST', '/import/upload', body, {
            'Content-Type': f'multipart/form-data; boundary={boundary}',
            'Content-Length': str(len(body)),
            'Cookie': self.cookie,
        })
        resp = conn.getresponse()
        html = resp.read().decode('utf-8')
        conn.close()

        self.assertEqual(resp.status, 200)
        self.assertIn('导入结果', html)
        self.assertIn('自动备份', html)
        self.assertIn('auto_before_import_', html)
        self.assertIn('新增房间', html)
        self.assertIn('新增业主', html)
        self.assertIn('未导入历史金额', html)
        self.assertIn('历史金额不会自动入账', html)
        self.assertNotIn('数据核对与修正', html)
        from server.db import get_db
        db = get_db()
        imported = db.execute("""
            SELECT r.room_number, r.contract_start, r.contract_end, o.name owner_name
            FROM rooms r LEFT JOIN owners o ON r.owner_id=o.id
            WHERE r.building='B座' AND r.room_number='901' LIMIT 1
        """).fetchone()
        db.close()
        self.assertIsNotNone(imported)
        self.assertEqual(imported['owner_name'], '自动备份业主')
        self.assertEqual(imported['contract_start'], '2024-01-01')
        self.assertEqual(imported['contract_end'], '2026-12-31')
        self.assertNotIn('/edit', html)
        self.assertIn('撤销本次导入', html)
        self.assertRegex(html, r'action="/backups/auto_before_import_[^"]+\.db/restore"')

        import server.db as db_module
        backups = os.listdir(db_module.BACKUP_DIR)
        self.assertTrue(any(name.startswith('auto_before_import_') for name in backups))


    def test_404_for_unknown_path(self):
        status, body = http_get('/this_does_not_exist', self.cookie, TEST_PORT)
        self.assertEqual(status, 404)


    def test_default_admin_password_warning_guides_local_first_run(self):
        status, body = http_get('/', self.cookie, TEST_PORT)
        self.assertEqual(status, 200)
        self.assertIn('默认管理员密码仍在使用', body)
        self.assertIn('/users/1/edit', body)
        self.assertIn('本地版不需要注册互联网账号', body)


