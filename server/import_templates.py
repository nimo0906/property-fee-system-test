#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Import template definitions and workbook rendering."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TEMPLATES = {
    'basic': {
        'sheet': '基础资料导入模板', 'filename': 'basic_info_template.xlsx',
        'headers': ['楼栋','单元/座','铺位号/房号','楼层','房间类型','面积㎡','物业费单价','水费标准','业主姓名','业主电话','身份证号','租户姓名','租户电话','租户身份证号','合同开始日期','合同结束日期','缴费周期','店铺名称','业态/商户类别','备注'],
        'rows': [['示例项目','商场','1F-101',1,'商户',88.5,4.8,'非居民','甲商贸','13900000000','','李四','13800000001','','2026-01-01','2026-12-31','季付','某某便利店','餐饮','基础资料模板，不填写历史收款金额'], ['B座','B座','902',9,'居民',95.2,'','非居民','王五','13800000000','','','','','2026-01-01','2026-12-31','月付','','','住宅示例，仅填写基础资料']],
        'notes': [['用途','导入房间、业主、租户、面积、合同期等基础资料。'], ['提醒','历史收款金额不要填在本模板，收款明细请用收款明细识别模板。']],
    },
    'owners': {
        'sheet': '业主信息模板', 'filename': 'owners_template.xlsx',
        'headers': ['业主姓名','业主电话','身份证号','备注'],
        'rows': [['张三','13800000000','', '只导入业主档案，不创建房间']],
        'notes': [['用途','只导入业主信息。'], ['必填','业主姓名。']],
    },
    'payment_ledger': {
        'sheet': '收款明细识别模板', 'filename': 'payment_ledger_template.xlsx',
        'headers': ['缴费日期','房间号','用户名称','本期金额','本期收款(合计)','物业费','水费','电费','垃圾费','其它费用'],
        'rows': [['2026-05-31','B座902','王五',500,500,300,100,80,20,0]],
        'notes': [['用途','只做收款明细识别和收费项目映射，不自动入账。'], ['必填','房间号、用户名称、本期金额、本期收款(合计)建议保留。']],
    },
    'bills': {
        'sheet': '账单记录模板', 'filename': 'bills_template.xlsx',
        'headers': ['楼栋','房号','费用项目','账期','金额','状态','已缴','缴费日期','缴费方式','备注'],
        'rows': [['B座','902','物业费(居民)','2026-05',180.88,'unpaid','','','','谨慎使用，建议先备份并预览']],
        'notes': [['用途','导入历史账单记录。'], ['提醒','账单导入会影响财务核对，正式使用前必须预览核对。']],
    },
    'b_tower_contracts': {
        'sheet': 'B座出租合同模板', 'filename': 'b_tower_contracts_template.xlsx',
        'headers': ['楼栋','单元/座','房号','租户','租户电话','面积','物业费单价','缴费周期','合同开始日期','合同结束日期','水费标准','备注'],
        'rows': [['B座','B座','902','李四','13800000001',95.2,1.9,'月付','2026-01-01','2026-12-31','非居民','默认只更新已有 B座房间']],
        'notes': [['用途','写入房间管理的租户、合同期、单价、周期等字段。'], ['提醒','默认不新建缺失房间，预览页可勾选允许新建。']],
    },
    'commercial_contracts': {
        'sheet': '在租合同', 'filename': 'commercial_contracts_template.xlsx',
        'headers': ['楼层','合同编号','商铺号','起租期（日）','承租人','进场日','免租装修期（日）','合同面积（㎡）','实际建筑面积（㎡）','租赁期限（年）','经营品牌','联系电话','交租方式','租金保证金','经营管理费/㎡元','租金/㎡元','物业费/㎡','垃圾费/㎡','水费/吨','电费/度','递增','备注'],
        'rows': [['一层','HT-2026-001','1F-101','2026-01-01','甲商贸','2026-01-10','30',88.5,90,1,'某某便利店','13800000001','季付',10000,20,100,5,1,5.8,0.85,'第二年递增3%','只读取“在租合同”sheet']],
        'notes': [['用途','写入合同档案和商业空间。'], ['必填','商铺号、合同编号、承租人、起租期、租赁期限、合同面积。'], ['Sheet','商业合同导入只读取“在租合同”sheet。']],
    },
}


def render_template_xlsx(key):
    spec = TEMPLATES.get(key)
    if not spec:
        return None, '', ''
    wb = Workbook()
    ws = wb.active
    ws.title = spec['sheet']
    ws.append(spec['headers'])
    for row in spec['rows']:
        ws.append(row)
    _style_sheet(ws, len(spec['headers']))
    note = wb.create_sheet('填写说明')
    note.append(['说明项', '内容'])
    for row in spec.get('notes', []):
        note.append(row)
    _style_sheet(note, 2, note_sheet=True)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), spec['filename'], spec['sheet']


def render_template_csv(key):
    spec = TEMPLATES.get(key)
    if not spec:
        return None, ''
    text = ','.join(spec['headers']) + '\n' + '\n'.join(','.join(str(x) for x in row) for row in spec['rows']) + '\n'
    return ('\ufeff' + text).encode('utf-8'), spec['filename'].replace('.xlsx', '.csv')


def _style_sheet(ws, width_count, note_sheet=False):
    fill = PatternFill('solid', fgColor='E6F4F1')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='0F3D38')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for idx in range(1, width_count + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 90 if note_sheet and idx == 2 else 16
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
