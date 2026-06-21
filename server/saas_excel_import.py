#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel import helpers for SaaS charge targets."""

import io

from fastapi import Depends, File, HTTPException, UploadFile
from fastapi.responses import HTMLResponse
from openpyxl import load_workbook

from server.saas_import_duplicate_preview import duplicate_preview_rows, existing_charge_targets, render_duplicate_preview_card
from server.saas_service import PermissionDenied
from server.saas_user_pages import _h, _page

VALID_HEADERS = [
    ('owner_name', '业主'), ('owner_phone', '联系电话'), ('building', '楼栋/区域'),
    ('unit', '单元/分区'), ('room_number', '房号/铺位号'), ('floor', '楼层'),
    ('shop_name', '店名'), ('tenant_name', '承租人'), ('tenant_phone', '承租电话'),
    ('category', '类型'), ('area', '面积'), ('unit_price_override', '独立单价'),
    ('payment_cycle', '缴费周期'), ('notes', '备注'),
]


def excel_preview_form():
    return '''<form method="post" action="/backoffice/imports/charge-targets/xlsx-preview" enctype="multipart/form-data"><label>Excel 文件预览</label><input name="xlsx_file" type="file" accept=".xlsx" required><button class="primary">上传 Excel 预览</button><div class="hint">仅支持 .xlsx；预览不会写库，确认后才写入有效行。</div></form>'''


def parse_xlsx_rows(content):
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or '').strip() for value in rows[0]]
    parsed = []
    for row in rows[1:]:
        if not any(value not in (None, '') for value in row):
            continue
        parsed.append({headers[idx]: '' if value is None else value for idx, value in enumerate(row) if idx < len(headers) and headers[idx]})
    return parsed


def _valid_header_cells():
    return ''.join(f'<th>{_h(label)}</th>' for _, label in VALID_HEADERS)


def _valid_row(row):
    cells = ''.join(f'<td>{_h(row.get(key))}</td>' for key, _ in VALID_HEADERS)
    return f'<tr>{cells}</tr>'


def _error_download_link(review):
    if not review.get('error_count'):
        return ''
    import_id = _h(review.get('import_id'))
    return f'<div class="actions" style="margin-bottom:12px"><a class="ghost-link" href="/api/imports/{import_id}/errors.csv">下载错误行 CSV</a></div><div class="hint">修正后可重新复制到导入预览；错误行下载不包含内部租户或项目编号。</div>'


def _render_excel_preview(user, review, duplicates=None):
    valid_rows = ''.join(_valid_row(row) for row in review['valid_rows']) or f'<tr><td colspan="{len(VALID_HEADERS)}">暂无有效行</td></tr>'
    error_rows = ''.join(f'<tr><td>{_h(err.get("row"))}</td><td>{_h((err.get("error") or "").replace("不能为空", "必填"))}</td></tr>' for err in review['errors']) or '<tr><td colspan="2">暂无错误</td></tr>'
    body = f'''
<section class="hero"><div><h1>导入预览</h1><div class="sub">有效 {review['valid_count']} 行，错误 {review['error_count']} 行。确认导入只会写入有效行。</div></div><div class="badge tenant-scope">{_h(user.get('tenant_name'))} · {_h(user.get('project_name'))}</div></section>
{render_duplicate_preview_card(duplicates or [])}
<section class="grid"><div class="card"><div class="card-h">有效行</div><div class="card-b"><table><thead><tr>{_valid_header_cells()}</tr></thead><tbody>{valid_rows}</tbody></table><form method="post" action="/backoffice/imports/charge-targets/confirm"><input type="hidden" name="import_id" value="{_h(review['import_id'])}"><button class="primary">确认导入</button></form></div></div>
<aside class="card"><div class="card-h">错误行</div><div class="card-b">{_error_download_link(review)}<table><thead><tr><th>行号</th><th>错误</th></tr></thead><tbody>{error_rows}</tbody></table></div></aside></section>'''
    return _page('导入预览', body)


def register_excel_import_routes(app, service, repository, current_user):
    @app.post('/backoffice/imports/charge-targets/xlsx-preview', response_class=HTMLResponse)
    async def preview_xlsx_import_page(xlsx_file: UploadFile = File(...), user=Depends(current_user)):
        try:
            if not str(xlsx_file.filename or '').lower().endswith('.xlsx'):
                raise HTTPException(status_code=400, detail='xlsx only')
            rows = parse_xlsx_rows(await xlsx_file.read())
            preview = service.preview_charge_target_import(user, user['project_id'], rows)
            review = service.get_import_review(user, user['project_id'], preview['import_id'])
            duplicates = duplicate_preview_rows(review['valid_rows'], existing_charge_targets(service, repository, user))
            return HTMLResponse(_render_excel_preview(user, review, duplicates))
        except HTTPException:
            raise
        except PermissionDenied:
            raise HTTPException(status_code=403, detail='forbidden')
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f'invalid xlsx: {exc}')
