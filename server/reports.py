#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Reports, statistics, and exports."""

import csv
import io
from datetime import date

import openpyxl
import openpyxl.styles

from server.base import BaseHandler
from server.db import add_months, get_db, get_period, h, m, qs, date_to_period, period_to_date


class ReportMixin(BaseHandler):

    def _report_period(self, q):
        raw = qs(q, 'period', get_period()).strip()
        if '~' in raw and len(raw) >= 15:
            return raw
        return date_to_period(raw)

    def _report_reconciliation_data(self, period, building='', status=''):
        db = get_db()
        cond = ['b.billing_period=?']
        vals = [period]
        if building:
            cond.append('r.building=?')
            vals.append(building)
        if status:
            if status == 'arrears':
                cond.append("b.status IN ('unpaid','overdue','partial')")
            else:
                cond.append('b.status=?')
                vals.append(status)
        where = ' AND '.join(cond)
        rows = db.execute(f'''SELECT b.*,r.building,r.unit,r.room_number,o.name owner_name,o.phone,f.name fee_name,
            COALESCE((SELECT SUM(amount_paid) FROM payments WHERE bill_id=b.id),0) paid
            FROM bills b JOIN rooms r ON b.room_id=r.id
            LEFT JOIN owners o ON b.owner_id=o.id LEFT JOIN fee_types f ON b.fee_type_id=f.id
            WHERE {where} ORDER BY r.building,r.room_number,b.id''', vals).fetchall()
        building_rows = db.execute(f'''SELECT r.building,COUNT(b.id) bill_count,COALESCE(SUM(b.amount),0) amount,
            COALESCE(SUM((SELECT COALESCE(SUM(amount_paid),0) FROM payments WHERE bill_id=b.id)),0) paid
            FROM bills b JOIN rooms r ON b.room_id=r.id
            WHERE {where} GROUP BY r.building ORDER BY r.building''', vals).fetchall()
        db.close()
        total = sum(float(r['amount'] or 0) for r in rows)
        paid = sum(float(r['paid'] or 0) for r in rows)
        due = total - paid
        rate = round(paid / total * 100, 1) if total else 0
        arrears = [r for r in rows if float(r['amount'] or 0) - float(r['paid'] or 0) > 0.001]
        return {'rows': rows, 'building_rows': building_rows, 'total': total, 'paid': paid, 'due': due, 'rate': rate, 'arrears': arrears}

    def _reports_reconciliation_csv(self, q):
        p = qs(q, 'period', get_period())
        building = qs(q, 'building')
        status = qs(q, 'status')
        data = self._report_reconciliation_data(p, building, status)
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(['building','room','owner','phone','fee_type','period','amount','paid','due','status','due_date','bill_number'])
        for r in data['rows']:
            due = float(r['amount'] or 0) - float(r['paid'] or 0)
            writer.writerow([r['building'] or '', r['room_number'] or '', r['owner_name'] or '', r['phone'] or '',
                             r['fee_name'] or '', r['billing_period'] or '', m(r['amount']), m(r['paid']), m(due),
                             r['status'] or '', r['due_date'] or '', r['bill_number'] or ''])
        raw = buf.getvalue().encode('utf-8-sig')
        self.send_response(200)
        self.send_header('Content-Type', 'text/csv; charset=utf-8')
        self.send_header('Content-Disposition', f'attachment; filename=reconciliation_{p}.csv')
        self.send_header('Content-Length', str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def _reports_collections_csv(self, q):
        p = qs(q, 'period', get_period())
        data = self._collection_summary_data(p)
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(['date','method','operator','fee_type','count','amount'])
        for r in data['rows']:
            w.writerow([r['day'], r['payment_method'], r['operator'], r['fee_name'], r['cnt'], m(r['amount'])])
        raw = buf.getvalue().encode('utf-8-sig')
        self.send_response(200)
        self.send_header('Content-Type', 'text/csv; charset=utf-8')
        self.send_header('Content-Disposition', f'attachment; filename=collections_{p}.csv')
        self.send_header('Content-Length', str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def _collection_summary_data(self, period):
        db = get_db()
        rows = db.execute('''SELECT date(p.payment_date) day,p.payment_method,COALESCE(p.operator,'') operator,f.name fee_name,
            COUNT(*) cnt,COALESCE(SUM(p.amount_paid),0) amount
            FROM payments p JOIN bills b ON p.bill_id=b.id LEFT JOIN fee_types f ON b.fee_type_id=f.id
            WHERE b.billing_period=? GROUP BY day,p.payment_method,p.operator,f.name
            ORDER BY day DESC,p.payment_method,p.operator,f.name''', (period,)).fetchall()
        totals = db.execute('''SELECT p.payment_method,COUNT(*) cnt,COALESCE(SUM(p.amount_paid),0) amount
            FROM payments p JOIN bills b ON p.bill_id=b.id WHERE b.billing_period=?
            GROUP BY p.payment_method ORDER BY p.payment_method''', (period,)).fetchall()
        db.close()
        return {'rows': rows, 'totals': totals, 'amount': sum(float(r['amount'] or 0) for r in rows)}

    def _arrears_analysis_data(self, period):
        db = get_db()
        by_fee = db.execute('''SELECT f.name fee_name,COUNT(b.id) cnt,COALESCE(SUM(b.amount-COALESCE((SELECT SUM(amount_paid) FROM payments WHERE bill_id=b.id),0)),0) due
            FROM bills b LEFT JOIN fee_types f ON b.fee_type_id=f.id
            WHERE b.billing_period=? AND b.status IN('unpaid','overdue','partial')
            GROUP BY f.name ORDER BY due DESC''', (period,)).fetchall()
        long_due = db.execute('''SELECT o.name owner_name,o.phone,r.building,r.room_number,COUNT(b.id) cnt,
            MIN(b.billing_period) first_period,COALESCE(SUM(b.amount-COALESCE((SELECT SUM(amount_paid) FROM payments WHERE bill_id=b.id),0)),0) due
            FROM bills b JOIN rooms r ON b.room_id=r.id LEFT JOIN owners o ON b.owner_id=o.id
            WHERE b.status IN('unpaid','overdue','partial')
            GROUP BY COALESCE(o.id,r.id) HAVING cnt>=2 OR first_period<?
            ORDER BY due DESC LIMIT 20''', (period,)).fetchall()
        db.close()
        return {'by_fee': by_fee, 'long_due': long_due}

    def _render_reconciliation_section(self, period, data, building='', status=''):
        building_rows = ''.join(
            f'''<tr><td>{h(r["building"])}</td><td class="text-end">{r["bill_count"]}</td>
            <td class="text-end"><span class="money">¥{m(r["amount"])}</span></td>
            <td class="text-end"><span class="money money-paid">¥{m(r["paid"])}</span></td>
            <td class="text-end"><span class="money money-due">¥{m(float(r["amount"] or 0)-float(r["paid"] or 0))}</span></td></tr>'''
            for r in data['building_rows']
        ) or '<tr><td colspan="5" class="text-center text-muted">暂无数据</td></tr>'
        arrears_rows = ''.join(
            f'''<tr><td>{h(r["bill_number"] or r["id"])}</td><td>{h(r["building"])}-{h(r["room_number"])}</td>
            <td>{h(r["owner_name"] or "未知")}</td><td>{h(r["phone"] or "")}</td><td>{h(r["fee_name"] or "")}</td>
            <td class="text-end"><span class="money">¥{m(r["amount"])}</span></td>
            <td class="text-end"><span class="money money-paid">¥{m(r["paid"])}</span></td>
            <td class="text-end"><span class="money money-due">¥{m(float(r["amount"] or 0)-float(r["paid"] or 0))}</span></td></tr>'''
            for r in data['arrears'][:50]
        ) or '<tr><td colspan="8" class="text-center text-muted">本账期无欠费</td></tr>'
        filter_query = f"period={h(period)}{'&amp;building='+h(building) if building else ''}{'&amp;status='+h(status) if status else ''}"
        collections = self._collection_summary_data(period)
        collection_rows = ''.join(
            f'''<tr><td>{h(r["day"])}</td><td>{h(r["payment_method"])}</td><td>{h(r["operator"] or "-")}</td><td>{h(r["fee_name"] or "-")}</td><td class="text-end">{r["cnt"]}</td><td class="text-end money money-paid">¥{m(r["amount"])}</td></tr>'''
            for r in collections['rows'][:80]
        ) or '<tr><td colspan="6" class="text-center text-muted">暂无收款</td></tr>'
        arrears = self._arrears_analysis_data(period)
        arrears_fee_rows = ''.join(f'<tr><td>{h(r["fee_name"] or "-")}</td><td class="text-end">{r["cnt"]}</td><td class="text-end money money-due">¥{m(r["due"])}</td></tr>' for r in arrears['by_fee']) or '<tr><td colspan="3" class="text-center text-muted">暂无欠费</td></tr>'
        business_rows_raw = []
        try:
            db = get_db()
            business_rows_raw = db.execute('''SELECT COALESCE(NULLIF(r.business_type,''),'未填写') business_type,COUNT(DISTINCT r.id) room_count,
                COALESCE(SUM(b.amount),0) amount,COALESCE(SUM((SELECT COALESCE(SUM(amount_paid),0) FROM payments WHERE bill_id=b.id)),0) paid
                FROM rooms r LEFT JOIN bills b ON b.room_id=r.id AND b.billing_period=?
                WHERE r.category='商户' OR COALESCE(r.business_type,'')<>''
                GROUP BY COALESCE(NULLIF(r.business_type,''),'未填写') ORDER BY amount DESC,room_count DESC''', (period,)).fetchall()
            db.close()
        except Exception:
            business_rows_raw = []
        business_rows = ''.join(
            f'<tr><td>{h(r["business_type"])}</td><td class="text-end">{r["room_count"]}</td><td class="text-end money">¥{m(r["amount"])}</td><td class="text-end money money-paid">¥{m(r["paid"])}</td></tr>'
            for r in business_rows_raw
        ) or '<tr><td colspan="4" class="text-center text-muted">暂无业态数据</td></tr>'
        long_due_rows = ''.join(
            f'<tr><td>{h(r["owner_name"] or "未知")}</td><td>{h(r["phone"] or "")}</td><td>{h(r["building"])}-{h(r["room_number"])}</td><td>{h(r["first_period"])}</td><td class="text-end">{r["cnt"]}</td><td class="text-end money money-due">¥{m(r["due"])}</td></tr>'
            for r in arrears['long_due']
        ) or '<tr><td colspan="6" class="text-center text-muted">暂无长期欠费</td></tr>'
        return f'''
        <div class="card mt-4"><div class="card-header d-flex justify-content-between align-items-center flex-wrap gap-2">
            <span>账期对账 <small class="text-muted">({h(period)})</small></span>
            <span class="export-actions">
                <a class="btn btn-sm btn-outline-secondary" href="/reports/reconciliation/print?{filter_query}"><i class="bi bi-printer"></i> 打印当前对账单</a>
                <a class="btn btn-sm btn-outline-secondary" href="/reports/reconciliation.csv?{filter_query}"><i class="bi bi-download"></i> 导出对账CSV</a>
            </span></div>
        <div class="card-body"><div class="row text-center g-2">
            <div class="col-md-3"><div class="summary-tile"><div class="label">应收合计</div><strong class="money">¥{m(data['total'])}</strong></div></div>
            <div class="col-md-3"><div class="summary-tile"><div class="label">已收合计</div><strong class="money money-paid">¥{m(data['paid'])}</strong></div></div>
            <div class="col-md-3"><div class="summary-tile"><div class="label">未收合计</div><strong class="money money-due">¥{m(data['due'])}</strong></div></div>
            <div class="col-md-3"><div class="summary-tile"><div class="label">收缴率</div><strong>{data['rate']}%</strong></div></div>
        </div></div></div>
        <div class="card"><div class="card-header">楼栋对账统计</div><div class="table-responsive"><table class="table table-sm mb-0">
        <thead><tr><th>楼栋</th><th class="text-end">账单数</th><th class="text-end">应收</th><th class="text-end">已收</th><th class="text-end">未收</th></tr></thead><tbody>{building_rows}</tbody></table></div></div>
        <div class="card"><div class="card-header">欠费清单</div><div class="table-responsive"><table class="table table-sm mb-0">
        <thead><tr><th>账单号</th><th>房间</th><th>业主</th><th>电话</th><th>项目</th><th class="text-end">应收</th><th class="text-end">已收</th><th class="text-end">未收</th></tr></thead><tbody>{arrears_rows}</tbody></table></div></div>
        <div class="card mt-3"><div class="card-header d-flex justify-content-between"><span>收款对账汇总</span><a class="btn btn-sm btn-outline-secondary" href="/reports/collections.csv?period={h(period)}">导出收款CSV</a></div><div class="table-responsive"><table class="table table-sm mb-0">
        <thead><tr><th>日期</th><th>方式</th><th>操作员</th><th>费用类型</th><th class="text-end">笔数</th><th class="text-end">金额</th></tr></thead><tbody>{collection_rows}</tbody></table></div></div>
        <div class="row g-3 mt-1"><div class="col-md-4"><div class="card"><div class="card-header">按费用类型欠费</div><table class="table table-sm mb-0"><thead><tr><th>项目</th><th class="text-end">笔数</th><th class="text-end">欠费</th></tr></thead><tbody>{arrears_fee_rows}</tbody></table></div></div>
        <div class="col-md-4"><div class="card"><div class="card-header">商户业态统计</div><table class="table table-sm mb-0"><thead><tr><th>业态</th><th class="text-end">房间</th><th class="text-end">应收</th><th class="text-end">实收</th></tr></thead><tbody>{business_rows}</tbody></table></div></div>
        <div class="col-md-4"><div class="card"><div class="card-header">长期欠费名单</div><table class="table table-sm mb-0"><thead><tr><th>业主</th><th>电话</th><th>房间</th><th>最早账期</th><th class="text-end">笔数</th><th class="text-end">欠费</th></tr></thead><tbody>{long_due_rows}</tbody></table></div></div></div>
        '''

    def _reports_reconciliation_print(self, q):
        period = qs(q, 'period', get_period())
        building = qs(q, 'building')
        status = qs(q, 'status')
        data = self._report_reconciliation_data(period, building, status)
        status_label = {'unpaid':'未缴','partial':'部分缴费','paid':'已缴','arrears':'全部欠费','':'全部状态'}.get(status, status)
        row_parts = []
        for r in data['rows']:
            due = float(r['amount'] or 0) - float(r['paid'] or 0)
            row_parts.append(
                f'<tr><td>{h(r["bill_number"] or r["id"])}</td><td>{h(r["building"])}-{h(r["room_number"])}</td>'
                f'<td>{h(r["owner_name"] or "未知")}</td><td>{h(r["phone"] or "")}</td><td>{h(r["fee_name"] or "")}</td>'
                f'<td class="amt">{m(r["amount"])}</td><td class="amt">{m(r["paid"])}</td><td class="amt">{m(due)}</td></tr>'
            )
        rows = ''.join(row_parts) or '<tr><td colspan="8" class="muted">暂无数据</td></tr>'
        back = f'/reports?period={h(period)}' + (f'&building={h(building)}' if building else '') + (f'&status={h(status)}' if status else '')
        html = f'''<!doctype html><html><head><meta charset="utf-8"><title>对账单打印</title>
        <style>
        body{{font-family:-apple-system,"PingFang SC","Microsoft YaHei",sans-serif;color:#111;margin:24px}}
        h1{{text-align:center;font-size:22px;margin:0 0 8px}} .meta{{text-align:center;color:#555;margin-bottom:18px}}
        .summary{{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:16px}}
        .box{{border:1px solid #ccc;padding:10px;text-align:center}} .box strong{{display:block;font-size:18px;margin-top:4px}}
        table{{width:100%;border-collapse:collapse;font-size:12px}} th,td{{border:1px solid #999;padding:6px}} th{{background:#f2f2f2}} .amt{{text-align:right}} .muted{{text-align:center;color:#777}}
        .actions{{margin-bottom:16px;text-align:right}} @media print{{.actions{{display:none}} body{{margin:0}}}}
        </style></head><body>
        <div class="actions"><button onclick="window.print()">打印</button> <a href="{back}">返回报表</a></div>
        <h1>对账单打印</h1><div class="meta">账期：{h(period)}　楼栋：{h(building or '全部')}　状态：{h(status_label)}</div>
        <div class="summary"><div class="box">应收合计<strong>{m(data['total'])}</strong></div><div class="box">已收合计<strong>{m(data['paid'])}</strong></div><div class="box">未收合计<strong>{m(data['due'])}</strong></div><div class="box">收缴率<strong>{data['rate']}%</strong></div></div>
        <table><thead><tr><th>账单号</th><th>房间</th><th>业主</th><th>电话</th><th>项目</th><th>应收</th><th>已收</th><th>未收</th></tr></thead><tbody>{rows}</tbody></table>
        </body></html>'''
        self._html(html)

    def _export_kingdee(self, q):
        p = qs(q, 'period', get_period())
        bld = qs(q, 'building')
        db = get_db()
        sql = "SELECT b.*,r.building,r.unit,r.room_number,r.category,o.name oname,"
        sql += "f.name ft_name,"
        sql += "COALESCE((SELECT SUM(amount_paid) FROM payments WHERE bill_id=b.id),0) paid "
        sql += "FROM bills b JOIN rooms r ON b.room_id=r.id "
        sql += "LEFT JOIN owners o ON b.owner_id=o.id "
        sql += "LEFT JOIN fee_types f ON b.fee_type_id=f.id "
        sql += "WHERE b.billing_period=?"
        vals = [p]
        if bld:
            sql += " AND r.building=?"
            vals.append(bld)
        sql += " ORDER BY r.building,r.room_number,f.name"
        rows = db.execute(sql, vals).fetchall()
        db.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kingdee_" + p
        hdr_font = openpyxl.styles.Font(bold=True)
        hdr_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_font_w = openpyxl.styles.Font(bold=True, color="FFFFFF")
        tb = openpyxl.styles.Border(left=openpyxl.styles.Side(style="thin"), right=openpyxl.styles.Side(style="thin"), top=openpyxl.styles.Side(style="thin"), bottom=openpyxl.styles.Side(style="thin"))
        ws.merge_cells("A1:J1")
        ws["A1"] = "Property Fee Details - " + p
        ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
        headers = ["Bill#","Building","Room","Owner","Fee Type","Amount","Paid","Due","Status","Due Date"]
        for ci, hh in enumerate(headers, 1):
            c = ws.cell(row=3, column=ci, value=hh)
            c.font = hdr_font_w; c.fill = hdr_fill; c.border = tb
        rn = 4; ta = 0; tp = 0; td = 0
        sm = {"paid":"Paid","unpaid":"Unpaid","overdue":"Overdue","partial":"Partial"}
        for r in rows:
            rm = r["amount"] - r["paid"]
            ws.cell(row=rn, column=1, value=r["bill_number"] or "").border = tb
            ws.cell(row=rn, column=2, value=r["building"] or "").border = tb
            ws.cell(row=rn, column=3, value=r["room_number"] or "").border = tb
            ws.cell(row=rn, column=4, value=r["oname"] or "N/A").border = tb
            ws.cell(row=rn, column=5, value=r["ft_name"] or "").border = tb
            ws.cell(row=rn, column=6, value=r["amount"]).border = tb
            ws.cell(row=rn, column=7, value=r["paid"]).border = tb
            ws.cell(row=rn, column=8, value=round(rm, 2)).border = tb
            ws.cell(row=rn, column=9, value=sm.get(r["status"], r["status"])).border = tb
            ws.cell(row=rn, column=10, value=r["due_date"] or "").border = tb
            ta += r["amount"]; tp += r["paid"]; td += rm; rn += 1
        for c in range(1, 11):
            ws.cell(row=rn, column=c).border = tb
            ws.cell(row=rn, column=c).font = hdr_font
        ws.cell(row=rn, column=1, value="TOTAL")
        ws.cell(row=rn, column=6, value=round(ta, 2))
        ws.cell(row=rn, column=7, value=round(tp, 2))
        ws.cell(row=rn, column=8, value=round(td, 2))
        for i, w in enumerate([18, 10, 10, 14, 18, 12, 12, 12, 10, 14], 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws2 = wb.create_sheet(title="Summary")
        ws2.merge_cells("A1:E1")
        ws2["A1"] = "Fee Summary - " + p
        ws2["A1"].font = openpyxl.styles.Font(bold=True, size=14)
        sh = ["Fee Type","Count","Total","Paid","Rate"]
        for ci, hh in enumerate(sh, 1):
            c = ws2.cell(row=3, column=ci, value=hh)
            c.font = hdr_font_w; c.fill = hdr_fill; c.border = tb
        fd = {}
        for r in rows:
            fn = r["ft_name"] or "Other"
            if fn not in fd: fd[fn] = {"cnt": 0, "total": 0, "paid": 0}
            fd[fn]["cnt"] += 1; fd[fn]["total"] += r["amount"]; fd[fn]["paid"] += r["paid"]
        rn2 = 4
        for fn, fv in fd.items():
            rate_str = str(round(fv["paid"] / fv["total"] * 100, 1)) + "%" if fv["total"] > 0 else "0%"
            ws2.cell(row=rn2, column=1, value=fn).border = tb
            ws2.cell(row=rn2, column=2, value=fv["cnt"]).border = tb
            ws2.cell(row=rn2, column=3, value=round(fv["total"], 2)).border = tb
            ws2.cell(row=rn2, column=4, value=round(fv["paid"], 2)).border = tb
            ws2.cell(row=rn2, column=5, value=rate_str).border = tb
            rn2 += 1
        for i, w in enumerate([20, 10, 14, 14, 10], 1):
            ws2.column_dimensions[chr(64 + i)].width = w
        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", "attachment; filename=kingdee_" + p + ".xlsx")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _reports(self, q):
        p=self._report_period(q);bld=qs(q,'building');db=get_db()
        fts=db.execute('''SELECT f.name,f.unit_price,f.calc_method,
            COUNT(b.id) cnt,COALESCE(SUM(b.amount),0) tot,
            COALESCE((SELECT SUM(p.amount_paid) FROM payments p WHERE p.bill_id IN (SELECT id FROM bills WHERE fee_type_id=f.id AND billing_period=?) ),0) paid
            FROM fee_types f LEFT JOIN bills b ON b.fee_type_id=f.id AND b.billing_period=?
            WHERE f.is_active=1 GROUP BY f.id ORDER BY f.sort_order''',(p,p)).fetchall()
        daily=db.execute('''SELECT DATE(p.payment_date) day,COUNT(*) cnt,SUM(p.amount_paid) tot
            FROM payments p JOIN bills b ON p.bill_id=b.id
            WHERE b.billing_period=? GROUP BY day ORDER BY day''',(p,)).fetchall()
        daily_rows=''.join(f'<tr><td>{h(r["day"])}</td><td class="text-end">{r["cnt"]}</td><td class="text-end"><span class="money money-paid">+¥{m(r["tot"])}</span></td></tr>' for r in daily)
        bld_stats=db.execute('''SELECT r.building,
            COUNT(DISTINCT b.id) bill_cnt,COALESCE(SUM(b.amount),0) tot,
            COALESCE((SELECT SUM(p.amount_paid) FROM payments p JOIN bills b2 ON p.bill_id=b2.id WHERE b2.room_id IN (SELECT id FROM rooms WHERE building=r.building) AND b2.billing_period=?),0) paid,
            COUNT(DISTINCT CASE WHEN b.status IN('unpaid','overdue','partial') THEN b.id END) unpaid_cnt,
            COUNT(DISTINCT r2.id) room_cnt
            FROM rooms r LEFT JOIN rooms r2 ON r2.building=r.building
            LEFT JOIN bills b ON b.room_id=r2.id AND b.billing_period=?
            GROUP BY r.building ORDER BY r.building''',(p,p)).fetchall()
        seen_bld=set();bld_rows_list=[]
        for r in bld_stats:
            if r['building'] not in seen_bld:
                seen_bld.add(r['building'])
                rate_val=round(r['paid']/r['tot']*100,1) if r['tot']>0 else 0
                bld_rows_list.append(f'''<tr><td><strong>{h(r["building"])}</strong></td>
                <td class="text-end">{r["room_cnt"]}</td><td class="text-end">{r["bill_cnt"]}</td>
                <td class="text-end"><span class="money">¥{m(r["tot"])}</span></td><td class="text-end"><span class="money money-paid">¥{m(r["paid"])}</span></td>
                <td class="text-end">{rate_val}%</td>
                <td class="text-end"><span class="badge status-{"danger" if r["unpaid_cnt"]>0 else "success"}">{r["unpaid_cnt"]}</span></td></tr>''')
        top10=db.execute('''SELECT o.id,o.name,o.phone,r.building,r.room_number,
            SUM(b.amount-COALESCE((SELECT SUM(amount_paid) FROM payments WHERE bill_id=b.id),0)) overdue
            FROM bills b JOIN rooms r ON b.room_id=r.id
            LEFT JOIN owners o ON b.owner_id=o.id
            WHERE b.billing_period=? AND b.status IN('unpaid','overdue','partial')
            GROUP BY o.id ORDER BY overdue DESC LIMIT 10''',(p,)).fetchall()
        top10_rows=''.join(f'<tr><td><strong>{h(r["name"]or"未知")}</strong></td><td>{h(r["building"]or"")}-{h(r["room_number"]or"")}</td><td class="text-end"><strong class="money money-due">¥{m(r["overdue"])}</strong></td></tr>' for r in top10)
        trends=[]
        for i in range(11,-1,-1):
            d=add_months(date.today(), -i)
            pp=f"{d.year}-{d.month:02d}"
            ta=db.execute("SELECT COALESCE(SUM(amount),0) FROM bills WHERE billing_period=?",(pp,)).fetchone()[0]
            tp=db.execute("SELECT COALESCE(SUM(p.amount_paid),0) FROM payments p JOIN bills b ON p.bill_id=b.id WHERE b.billing_period=?",(pp,)).fetchone()[0]
            r=round(tp/ta*100,1) if ta>0 else 0
            trends.append({'period':pp,'label':f"{d.year}年{d.month}月",'total':ta,'paid':tp,'rate':r})
        waiver=db.execute("SELECT COALESCE(SUM(a.old_amount-a.new_amount),0) waiver_total,COUNT(*) waiver_cnt FROM bill_adjustments a JOIN bills b ON a.bill_id=b.id WHERE b.billing_period=? AND a.new_amount<a.old_amount",(p,)).fetchone()
        waiver_amt=waiver['waiver_total'] if waiver else 0;waiver_cnt=waiver['waiver_cnt'] if waiver else 0
        blds=db.execute("SELECT DISTINCT building FROM rooms ORDER BY building").fetchall()
        periods=db.execute("SELECT DISTINCT billing_period FROM bills ORDER BY billing_period DESC").fetchall()
        db.close()
        ta=sum(f['tot'] for f in fts);tp=sum(f['paid'] for f in fts);rr=round(tp/ta*100,1) if ta>0 else 0
        fr=''.join(f'''<tr><td><span class="badge status-neutral">{h(f["name"])}</span></td><td class="text-end">{f["cnt"]}</td>
    <td class="text-end"><span class="money">¥{m(f["tot"])}</span></td><td class="text-end"><span class="money money-paid">¥{m(f["paid"])}</span></td>
    <td class="text-end">{"<span class='text-success'>"+str(round(f["paid"]/f["tot"]*100,1) if f["tot"]>0 else 0)+"%</span>"}</td></tr>''' for f in fts)
        trd=''.join(f'''<tr><td>{t["label"]}</td><td class="text-end"><span class="money">¥{m(t["total"])}</span></td><td class="text-end"><span class="money money-paid">¥{m(t["paid"])}</span></td>
    <td class="text-end">{"<span class='text-success'>"+str(t["rate"])+"%</span>" if t["rate"]>=80 else "<span class='text-warning'>"+str(t["rate"])+"%</span>"}</td>
    <td style="min-width:120px"><div class="progress progress-thin" style="height:12px"><div class="progress-bar bg-success" style="width:{t["rate"]}%"></div></div></td>
    <td><div style="display:flex;align-items:end;height:40px;gap:1px"><div style="width:16px;height:{max(4,t["rate"]/100*36):.0f}px;background:{"#087443" if t["rate"]>=80 else "#b76e00"};border-radius:2px 2px 0 0"></div></div></td></tr>''' for t in trends)
        bld_opts='<option value="">全部楼栋</option>'+''.join(f'<option value="{h(b["building"])}"{" selected" if bld==b["building"] else""}>{h(b["building"])}</option>' for b in blds)
        period_values=[r['billing_period'] for r in periods]
        if p not in period_values:
            period_values.insert(0, p)
        period_opts=''.join(f'<option value="{h(pp)}"{" selected" if p==pp else ""}>{h(pp)}</option>' for pp in period_values)
        self._html(self._page('对账报表',f'''
    {self._render_reconciliation_section(p, self._report_reconciliation_data(p, bld, qs(q, 'status')), bld, qs(q, 'status'))}
    <div class="filter-bar">
    <form class="row g-2 align-items-end" method=GET>
    <div class="col-auto"><label class="form-label small text-muted mb-1">账期</label><select name="period" class="form-select form-select-sm" onchange="this.form.submit()">{period_opts}</select></div>
    <div class="col-auto"><label class="form-label small text-muted mb-1">楼栋</label><select name="building" class="form-select form-select-sm" onchange="this.form.submit()">{bld_opts}</select></div>
    <div class="col-auto"><label class="form-label small text-muted mb-1">状态</label><select name="status" class="form-select form-select-sm" onchange="this.form.submit()">
    <option value="">全部状态</option><option value="unpaid"{" selected" if qs(q,'status')=='unpaid' else ""}>未缴</option><option value="partial"{" selected" if qs(q,'status')=='partial' else ""}>部分缴费</option><option value="paid"{" selected" if qs(q,'status')=='paid' else ""}>已缴</option><option value="arrears"{" selected" if qs(q,'status')=='arrears' else ""}>全部欠费</option></select></div>
    <div class="col-auto"><button class="btn btn-sm btn-outline-primary"><i class="bi bi-search"></i> 筛选</button>
    <a href="/reports" class="btn btn-sm btn-outline-secondary"><i class="bi bi-x-circle"></i></a></div></form>
    </div>
    <div class="metric-grid">
    <div class="metric-card primary"><div class="metric-label">应收总额</div><div class="metric-value money">¥{m(ta)}</div></div>
    <div class="metric-card success"><div class="metric-label">已收总额</div><div class="metric-value money money-paid">¥{m(tp)}</div></div>
    <div class="metric-card {"success" if rr>=80 else "warning"}"><div class="metric-label">收费率</div><div class="metric-value {"text-success" if rr>=80 else "text-warning"}">{rr}%</div></div>
    <div class="metric-card danger"><div class="metric-label">减免金额</div><div class="metric-value money {"money-due" if waiver_amt>0 else "money-muted"}">¥{m(waiver_amt)}</div><small class="text-muted">{waiver_cnt}笔</small></div>
    </div>
    <div class="row g-4">
    <div class="col-md-6"><div class="card"><div class="card-header">各费用类型统计 <small class="text-muted">({p})</small></div>
    <div class="card-body p-0"><div class="table-responsive"><table class="table table-hover mb-0"><thead><tr><th>类型</th><th class="text-end">户数</th><th class="text-end">应收</th><th class="text-end">实收</th><th class="text-end">率</th></tr></thead>
    <tbody>{fr or '<tr><td colspan="5" class="text-center text-muted py-3">暂无数据</td></tr>'}</tbody></table></div></div></div></div>
    <div class="col-md-6"><div class="card"><div class="card-header">收款日报 <small class="text-muted">({p})</small></div>
    <div class="card-body p-0"><div class="table-responsive"><table class="table table-hover mb-0"><thead><tr><th>日期</th><th class="text-end">笔数</th><th class="text-end">金额</th></tr></thead>
    <tbody>{daily_rows or '<tr><td colspan="3" class="text-center text-muted py-3">暂无收款记录</td></tr>'}</tbody></table></div></div></div></div>
    </div>
    <div class="row g-4 mt-1">
    <div class="col-md-6"><div class="card"><div class="card-header">楼栋缴费统计 <small class="text-muted">({p})</small></div>
    <div class="card-body p-0"><div class="table-responsive"><table class="table table-hover mb-0"><thead><tr><th>楼栋</th><th class="text-end">房间</th><th class="text-end">笔数</th><th class="text-end">应收</th><th class="text-end">实收</th><th class="text-end">率</th><th class="text-end">欠费</th></tr></thead>
    <tbody>{''.join(bld_rows_list) or '<tr><td colspan="7" class="text-center text-muted py-3">暂无数据</td></tr>'}</tbody></table></div></div></div></div>
    <div class="col-md-6"><div class="card"><div class="card-header"><i class="bi bi-exclamation-triangle text-danger"></i> 欠费排行Top10 <small class="text-muted">({p})</small></div>
    <div class="card-body p-0"><div class="table-responsive"><table class="table table-hover mb-0"><thead><tr><th>业主</th><th>房间</th><th class="text-end">欠费金额</th></tr></thead>
    <tbody>{top10_rows or '<tr><td colspan="3" class="text-center text-muted py-3">本账期无欠费</td></tr>'}</tbody></table></div></div></div></div>
    </div>
    <div class="card mt-4"><div class="card-header">近12个月收费趋势</div>
    <div class="card-body"><canvas id="trendChart" height="100"></canvas></div>
    <div class="card-body p-0"><div class="table-responsive"><table class="table table-hover mb-0"><thead><tr><th>月份</th><th class="text-end">应收</th><th class="text-end">实收</th><th class="text-end">率</th><th>进度条</th><th>柱状图</th></tr></thead>
    <tbody>{trd or '<tr><td colspan="6" class="text-center text-muted py-3">暂无</td></tr>'}</tbody></table></div></div></div>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.5.0/dist/chart.umd.min.js"></script>
    <script>
    var _labels = [''' + '","'.join('"{t["label"]}"' for t in trends) + '''];
    var _totals = [''' + '","'.join('{m(t["total"])}' for t in trends) + '''];
    var _paids = [''' + '","'.join('{m(t["paid"])}' for t in trends) + '''];
    var ctx = document.getElementById('trendChart');
    if(ctx && typeof Chart !== 'undefined'){
    new Chart(ctx, {type:'bar',data:{labels:_labels,datasets:[
    {label:'应收',data:_totals,backgroundColor:'rgba(8,112,99,0.42)'},
    {label:'实收',data:_paids,backgroundColor:'rgba(8,116,67,0.68)'}
    ]},options:{responsive:true,scales:{y:{beginAtZero:true}}}});
    }
    </script>''','reports'))
