#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Enterprise analysis XLSX export."""

import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _style_header(row):
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws):
    for col in ws.columns:
        width = 12
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 3, 28))
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


def _append_table(ws, headers, rows):
    ws.append(headers)
    _style_header(ws[1])
    for row in rows:
        ws.append(row)
    _autosize(ws)


def _overview_sheet(wb, metrics):
    ws = wb.active
    ws.title = "经营概览"
    ws["A1"] = "经营分析看板"
    ws["A1"].font = Font(bold=True, size=16)
    rows = [
        ("统计账期", metrics.get("period", "")),
        ("应收合计", metrics.get("total_due", 0)),
        ("实收合计", metrics.get("total_paid", 0)),
        ("未收合计", metrics.get("unpaid_amount", 0)),
        ("收缴率", f'{metrics.get("collection_rate", 0)}%'),
    ]
    for idx, row in enumerate(rows, 2):
        ws.cell(idx, 1, row[0])
        ws.cell(idx, 2, row[1])
    _autosize(ws)


def _comparison_sheet(wb, metrics):
    ws = wb.create_sheet("B座商场对比")
    segments = metrics.get("segments", {})
    rows = []
    for name in ("B座", "商场"):
        item = segments.get(name, {})
        rows.append([
            name, item.get("due", 0), item.get("paid", 0),
            item.get("unpaid", 0), f'{item.get("collection_rate", 0)}%',
            item.get("bill_count", 0),
        ])
    _append_table(ws, ["范围", "应收", "已收", "未收", "收缴率", "账单数"], rows)


def _arrears_sheet(wb, metrics):
    ws = wb.create_sheet("欠费趋势")
    rows = [
        [item.get("period"), item.get("due", 0), item.get("paid", 0), item.get("unpaid", 0)]
        for item in metrics.get("arrears_trend", [])
    ]
    _append_table(ws, ["月份", "应收", "已收", "欠费"], rows)


def _merchant_sheet(wb, metrics):
    ws = wb.create_sheet("商户贡献")
    rows = [
        [item.get("merchant"), item.get("due", 0), item.get("paid", 0), item.get("unpaid", 0), item.get("bill_count", 0)]
        for item in metrics.get("merchant_contribution", [])
    ]
    _append_table(ws, ["商户/店铺", "应收", "已收", "未收", "账单数"], rows)


def _fee_sheet(wb, metrics):
    ws = wb.create_sheet("费用结构")
    rows = [
        [item.get("fee_name"), item.get("due", 0), item.get("bill_count", 0)]
        for item in metrics.get("fee_structure", [])
    ]
    _append_table(ws, ["收费项目", "应收", "账单数"], rows)


def build_enterprise_analysis_xlsx(metrics):
    wb = openpyxl.Workbook()
    _overview_sheet(wb, metrics)
    _comparison_sheet(wb, metrics)
    _arrears_sheet(wb, metrics)
    _merchant_sheet(wb, metrics)
    _fee_sheet(wb, metrics)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
