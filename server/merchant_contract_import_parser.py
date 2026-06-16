#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Parse rented-contract Excel rows for contract archive import."""

import io
import re
from datetime import datetime, timedelta

import openpyxl

from server.db import add_months

MAX_UPLOAD_SIZE = 10 * 1024 * 1024
REQUIRED_SHEET = "在租合同"
HEADER_NAMES = {
    "floor": "楼层",
    "contract_no": "合同编号",
    "space_no": "商铺号",
    "start_date": "起租期（日）",
    "merchant_name": "承租人",
    "entry_date": "进场日",
    "free_period": "免租装修期（日）",
    "contract_area": "合同面积（㎡）",
    "building_area": "实际建筑面积（㎡）",
    "lease_term": "租赁期限（年）",
    "shop_name": "经营品牌",
    "phone": "联系电话",
    "rent_cycle": "交租方式",
    "deposit_amount": "租金保证金",
    "management_rate": "经营管理费/㎡元",
    "rent_rate": "租金/㎡元",
    "property_rate": "物业费/㎡",
    "garbage_rate": "垃圾费/㎡",
    "water_rate": "水费/吨",
    "electricity_rate": "电费/度",
    "increase_rule": "递增",
    "notes": "备注",
}
CYCLE_MAP = {"月付": "monthly", "季付": "quarterly", "半年付": "semiannual", "年付": "yearly"}


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def _num(value):
    text = _clean(value).replace(",", "")
    if not text or text in ("/", "-", "否", "是"):
        return 0.0
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else 0.0


def _is_non_area_rent(value):
    text = _clean(value)
    return bool(text and any(key in text for key in ("%", "％", "销售额", "提成", "保底")))


def _rent_rate(value):
    return 0.0 if _is_non_area_rent(value) else _num(value)


def _date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = _clean(value)
    if not text:
        return ""
    text = text.replace("/", "-").replace(".", "-")
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if not match:
        return ""
    y, mo, d = [int(x) for x in match.groups()]
    try:
        return datetime(y, mo, d).date().isoformat()
    except ValueError:
        return ""


def _end_date(start_date, lease_term):
    if not start_date:
        return ""
    months = _lease_months(lease_term)
    if months <= 0:
        return ""
    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    return (add_months(start, months) - timedelta(days=1)).isoformat()


def _lease_months(value):
    text = _clean(value)
    if not text:
        return 0
    years = re.search(r"(\d+(?:\.\d+)?)\s*年", text)
    months = re.search(r"(\d+(?:\.\d+)?)\s*个?月", text)
    days = re.search(r"(\d+)\s*天", text)
    total = 0
    if years:
        total += int(round(float(years.group(1)) * 12))
    if months:
        total += int(round(float(months.group(1))))
    if days and not total:
        total += 1
    if total:
        return total
    try:
        return int(round(float(text) * 12))
    except ValueError:
        return 0


def _cycle(value):
    text = _clean(value)
    for key, cycle in CYCLE_MAP.items():
        if key in text:
            return cycle
    if "交6" in text or "半年" in text:
        return "semiannual"
    if "一次" in text or "年" in text:
        return "yearly"
    return "monthly"


def _floor_no(value):
    text = _clean(value)
    cn = {"负二层": -2, "负一层": -1, "一层": 1, "二层": 2, "三层": 3, "四层": 4, "五层": 5, "六层": 6, "七层": 7}
    if text in cn:
        return cn[text]
    match = re.search(r"-?\d+", text)
    return int(match.group(0)) if match else 1


def _load_rows(filename, raw_data):
    if not filename.lower().endswith(".xlsx"):
        raise ValueError("合同档案导入仅支持 .xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(raw_data), data_only=True, read_only=True)
    if REQUIRED_SHEET not in wb.sheetnames:
        raise ValueError("未找到“在租合同”sheet")
    ws = wb[REQUIRED_SHEET]
    rows = [[_clean(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    header_idx = -1
    for idx, row in enumerate(rows[:20]):
        if "商铺号" in row and "合同编号" in row and "承租人" in row:
            header_idx = idx
            break
    if header_idx < 0:
        raise ValueError("未识别到在租合同表头")
    headers = rows[header_idx]
    positions = {key: headers.index(label) for key, label in HEADER_NAMES.items() if label in headers}
    parsed = []
    for raw_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        def val(key):
            idx = positions.get(key)
            return row[idx] if idx is not None and idx < len(row) else ""
        space_no = val("space_no")
        contract_no = val("contract_no")
        merchant = val("merchant_name")
        if not space_no or space_no == "合计" or not merchant:
            continue
        start = _date(val("start_date"))
        lease_term = val("lease_term")
        contract_area = _num(val("contract_area"))
        building_area = _num(val("building_area"))
        rent_raw = val("rent_rate")
        rent_rate = _rent_rate(rent_raw)
        parsed.append({
            "row_no": raw_idx,
            "floor": _floor_no(val("floor")),
            "contract_no": contract_no or f"HT-{space_no}",
            "space_no": space_no,
            "start_date": start,
            "end_date": _end_date(start, lease_term),
            "merchant_name": merchant,
            "entry_date": _date(val("entry_date")),
            "free_period": val("free_period"),
            "contract_area": contract_area,
            "building_area": building_area,
            "lease_term": lease_term,
            "shop_name": val("shop_name") or merchant,
            "phone": val("phone"),
            "rent_cycle": _cycle(val("rent_cycle")),
            "deposit_amount": _num(val("deposit_amount")),
            "management_rate": _num(val("management_rate")),
            "rent_rate": rent_rate,
            "rent_raw": rent_raw,
            "rent_mode": special_rent_mode(rent_raw),
            "turnover_rate": special_rent_rate(rent_raw),
            "property_rate": _num(val("property_rate")),
            "garbage_rate": _num(val("garbage_rate")),
            "water_rate": _num(val("water_rate")),
            "electricity_rate": _num(val("electricity_rate")),
            "increase_rule": val("increase_rule"),
            "notes": val("notes"),
            "errors": _row_errors(space_no, contract_no, merchant, start, lease_term, contract_area, rent_rate, rent_raw),
        })
    return parsed


def _row_errors(space_no, contract_no, merchant, start, lease_term, area, rent_rate, rent_raw=""):
    errors = []
    if not space_no:
        errors.append("商铺号为空")
    if not contract_no:
        errors.append("合同编号为空，将用商铺号生成")
    if not merchant:
        errors.append("承租人为空")
    if not start:
        errors.append("起租期无法识别")
    if _lease_months(lease_term) <= 0:
        errors.append("租赁期限无法识别")
    if area <= 0:
        errors.append("合同面积异常")
    if _is_non_area_rent(rent_raw):
        errors.append("百分比/销售额租金需人工核对")
    elif rent_rate <= 0:
        errors.append("租金单价为空或异常")
    return errors


def special_rent_mode(value):
    from server.special_rent import detect_special_rent_mode
    return detect_special_rent_mode(value)


def special_rent_rate(value):
    from server.special_rent import extract_turnover_rate
    return extract_turnover_rate(value)
