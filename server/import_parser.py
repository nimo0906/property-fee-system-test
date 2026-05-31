#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Spreadsheet parsing helpers for data import."""

import csv
import io
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta

import openpyxl


SUPPORTED_EXTENSIONS = {'.csv', '.xlsx', '.xls'}
NAMESPACES = {
    'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
}


def parse_rows(filename, raw_data):
    ext = os.path.splitext(filename)[1].lower()
    if ext == '.csv':
        return _parse_csv(raw_data)
    if ext == '.xlsx':
        return _parse_xlsx(raw_data)
    if ext == '.xls':
        return _parse_xls(raw_data)
    raise ValueError(f'不支持的文件格式 {ext}')


def detect_header_row(rows, column_map, scan_limit=20):
    best_index = 0
    best_score = -1
    for i, row in enumerate(rows[:scan_limit]):
        score = 0
        for cell in row:
            key = normalize_header(cell)
            if key in column_map:
                score += 1
        if score > best_score:
            best_index = i
            best_score = score
    return best_index


def normalize_header(value):
    return str(value or '').strip().replace('\n', '').replace('\r', '').lower()


def clean_cell(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace('\r', '').replace('\n', ' ').strip()


def detect_data_type(col_map, headers):
    header_text = set(normalize_header(h) for h in headers)
    ledger_markers = {'缴费日期', '房间号', '本期收款(合计)', '本期金额'}
    if len(ledger_markers.intersection(header_text)) >= 3:
        return 'payment_ledger'
    if 'building' in col_map or 'room_number' in col_map:
        if 'amount' in col_map or 'period' in col_map:
            return 'bills' if 'amount' in col_map else 'rooms'
        return 'rooms'
    if 'owner_name' in col_map or 'owner_phone' in col_map:
        return 'owners'
    if 'fee_type_name' in col_map and 'amount' in col_map:
        return 'bills'
    return 'unknown'


def build_column_map(headers, column_map):
    result = {}
    for i, header in enumerate(headers):
        key = normalize_header(header)
        matched = column_map.get(key) or column_map.get(str(header or '').strip())
        if matched:
            result[matched] = i
    return result


def enrich_column_map_from_subheader(col_map, headers, rows, column_map, header_idx):
    if header_idx + 1 >= len(rows):
        return col_map
    result = dict(col_map)
    subheaders = rows[header_idx + 1]
    for i, value in enumerate(subheaders):
        key = normalize_header(value)
        matched = column_map.get(key) or column_map.get(str(value or '').strip())
        if matched and matched not in result:
            result[matched] = i
    return result


def summarize_rows(rows, headers, max_rows=8):
    sample = []
    width = min(len(headers), 12)
    for row in rows[:max_rows]:
        padded = row + [''] * max(0, width - len(row))
        sample.append(padded[:width])
    return sample


def normalize_date(value):
    text = str(value or '').strip()
    if not text:
        return ''
    text = text.replace('/', '.').replace('年', '.').replace('月', '.').replace('日', '')
    match = re.search(r'(\d{4})\.(\d{1,2})\.(\d{1,2})', text)
    if not match:
        match = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', text)
    if not match:
        return ''
    y, m, d = [int(x) for x in match.groups()]
    try:
        return datetime(y, m, d).strftime('%Y-%m-%d')
    except ValueError:
        return ''


def split_contract_period(value):
    text = str(value or '').strip()
    if not text:
        return '', ''
    parts = re.split(r'\s*[-~—至到]\s*', text, maxsplit=1)
    if len(parts) < 2:
        return normalize_date(text), ''
    return normalize_date(parts[0]), normalize_date(parts[1])


def infer_floor(room_number):
    text = str(room_number or '').upper()
    match = re.search(r'(\d{3,5})', text)
    if not match:
        return 1
    digits = match.group(1)
    return int(digits[:-2] or '1')


def infer_building(room_number, default='B座'):
    text = str(room_number or '').upper()
    match = re.search(r'([A-Z])座', text)
    if match:
        return match.group(1) + '座'
    if 'B座' in room_number or text.startswith('B'):
        return 'B座'
    if 'A座' in room_number or text.startswith('A'):
        return 'A座'
    return default


def clean_room_number(value):
    text = str(value or '').strip()
    text = text.replace('商用-', '').replace('住宅-', '').replace('写字楼-', '').strip()
    text = re.sub(r'^[AB]座', '', text, flags=re.I)
    text = re.sub(r'^B(?=\d{3,5}$)', '', text, flags=re.I)
    return text or str(value or '').strip()


def _parse_csv(raw_data):
    text = raw_data.decode('utf-8-sig')
    rows = []
    for row in csv.reader(io.StringIO(text)):
        cleaned = [clean_cell(c) for c in row]
        if any(cleaned):
            rows.append(cleaned)
    return rows


def _parse_xlsx(raw_data):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_data), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            cleaned = [clean_cell(v) for v in row]
            if any(cleaned):
                rows.append(cleaned)
        wb.close()
        return rows
    except Exception:
        return _parse_xlsx_raw(raw_data)


def _parse_xlsx_raw(raw_data):
    with zipfile.ZipFile(io.BytesIO(raw_data)) as zf:
        shared = _read_shared_strings(zf)
        sheet_path = _first_sheet_path(zf)
        root = ET.fromstring(zf.read(sheet_path))
    rows = []
    for row_el in root.findall('.//a:sheetData/a:row', NAMESPACES):
        values = []
        for cell in row_el.findall('a:c', NAMESPACES):
            ref = cell.attrib.get('r', '')
            col_index = _column_index(ref)
            while len(values) < col_index:
                values.append('')
            values.append(_cell_value(cell, shared))
        if any(values):
            rows.append(values)
    return rows


def _parse_xls(raw_data):
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError('缺少 xlrd，无法读取 .xls。请先安装 requirements.txt 里的依赖。') from exc
    book = xlrd.open_workbook(file_contents=raw_data)
    sheet = book.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        values = [clean_cell(_xls_cell_value(sheet.cell(r, c), book)) for c in range(sheet.ncols)]
        if any(values):
            rows.append(values)
    return rows


def _xls_cell_value(cell, book):
    try:
        import xlrd
    except ImportError:
        return cell.value
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            dt = datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode))
            return dt.strftime('%Y-%m-%d')
        except Exception:
            return cell.value
    return cell.value


def _read_shared_strings(zf):
    if 'xl/sharedStrings.xml' not in zf.namelist():
        return []
    root = ET.fromstring(zf.read('xl/sharedStrings.xml'))
    strings = []
    for si in root.findall('a:si', NAMESPACES):
        parts = [t.text or '' for t in si.findall('.//a:t', NAMESPACES)]
        strings.append(''.join(parts))
    return strings


def _first_sheet_path(zf):
    workbook = ET.fromstring(zf.read('xl/workbook.xml'))
    rels = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
    first_sheet = workbook.find('.//a:sheets/a:sheet', NAMESPACES)
    rel_id = first_sheet.attrib.get('{%s}id' % NAMESPACES['r'])
    for rel in rels:
        if rel.attrib.get('Id') == rel_id:
            target = rel.attrib.get('Target', 'worksheets/sheet1.xml')
            return 'xl/' + target.lstrip('/')
    return 'xl/worksheets/sheet1.xml'


def _cell_value(cell, shared):
    value_el = cell.find('a:v', NAMESPACES)
    if value_el is None:
        inline = cell.find('.//a:t', NAMESPACES)
        return clean_cell(inline.text if inline is not None else '')
    raw = value_el.text or ''
    if cell.attrib.get('t') == 's':
        try:
            return clean_cell(shared[int(raw)])
        except Exception:
            return ''
    return clean_cell(raw)


def _column_index(ref):
    match = re.match(r'([A-Z]+)', ref or '')
    if not match:
        return 1
    total = 0
    for char in match.group(1):
        total = total * 26 + ord(char) - ord('A') + 1
    return total
